def sub_proc1():
    import cx_Oracle as ora
    import pymysql
    import pandas as pd
    import os
    import glob
    import sys
    from datetime import datetime, timedelta
    import csv
    from email.message import EmailMessage
    import smtplib
    import os.path
    import ssl
    import numpy as np
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side, colors
    from openpyxl.formatting.rule import Rule
    from openpyxl.drawing.image import Image
    import xlsxwriter
    #from passwd import pswd
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
   
    from cryptography.fernet import Fernet
    import configparser
    from collections import OrderedDict
    
    os.chdir("C:\Parts POSB Automation")
    
    # Load the encryption key
    with open('encryption_key.key', 'rb') as file:
        key = file.read()
    
    # Create a Fernet cipher using the key
    cipher_suite = Fernet(key)
    
    # Read the encrypted data from the encrypted_config.ini file
    with open('encrypted_config.ini', 'rb') as file:
        ciphertext = file.read()
    
    # Decrypt the ciphertext using the cipher
    plaintext = cipher_suite.decrypt(ciphertext)
    
    # Convert the plaintext to a string
    plaintext_str = plaintext.decode('utf-8')
    
    # Parse the decrypted config data into a dictionary
    config_data = configparser.ConfigParser(dict_type=OrderedDict)
    config_data.read_string(plaintext_str)
    
    credentials = config_data['Credentials']
    connstr_plsql_v = credentials['connstr_plsql']
    connstr_mysql_v = credentials['connstr_mysql']
    host_mysql_v = credentials['host_mysql']
    user_mysql_v = credentials['user_mysql']
    pswd_mysql_v = credentials['pswd_mysql']
    dbas_mysql_v = credentials['dbas_mysql']
    passwd_email = credentials['passwd_email']
    sender_email = credentials['sender_email']
    receiv_email = credentials['receiv_email']
    cc_email = credentials['cc_email']
    attention = credentials['attention0']
    working_folder = credentials['working_folder']
    os.chdir(working_folder)

    print(working_folder)

    
    conn=pymysql.connect(host=host_mysql_v, user=user_mysql_v,  passwd=pswd_mysql_v, database=dbas_mysql_v)
    
    ora.init_oracle_client(
        lib_dir=r"C:\cx_Oracle\instantclient-basic-windows.x64-21.8.0.0.0dbru\instantclient-basic-windows.x64-21.8.0.0.0dbru\instantclient_21_8")
    #connstr = 'RO999999999/xxxxxxxx@geaebsdbadg1.appl.ge.com:1521/ercebs1p'
    connection = ora.connect(connstr_plsql_v)
    # cur=connection.cursor()
    now = datetime.now()
    if now.weekday() == 0:
        # If it is Monday, set unstruct_fromDate to two days ago
        unstruct_fromDate = now + timedelta(days=-3)
        unstruct_fromDate3 = now + timedelta(days=-10)                              
        fromDate3 = unstruct_fromDate3.strftime("%d-%b-%Y 00:00:01").upper()
        print("date and time =", fromDate3)
        unstruct_toDate3 = now + timedelta(days=-4)                              
        toDate3 = unstruct_toDate3.strftime("%d-%b-%Y 23:59:59").upper()
    else:
        # If it is not Monday, set unstruct_fromDate to one day ago
        unstruct_fromDate = now + timedelta(days=-1)
        unstruct_fromDate3 = now + timedelta(days=-13)                              
        fromDate3 = unstruct_fromDate3.strftime("%d-%b-%Y 00:00:01").upper()
        print("date and time =", fromDate3)
        unstruct_toDate3 = now + timedelta(days=-2)                              
        toDate3 = unstruct_toDate3.strftime("%d-%b-%Y 23:59:59").upper()

    unstruct_toDate = now + timedelta(days=-1)           
    fromDate = unstruct_fromDate.strftime("%d-%b-%Y 00:00:01").upper()
    fromDate2 = unstruct_fromDate.strftime("%Y-%m-%d 00:00:01").upper()
    print("date and time =", fromDate)
    print("date and time =", fromDate2)
    toDate = unstruct_toDate.strftime("%d-%b-%Y 23:59:59").upper()
    toDate2 = unstruct_toDate.strftime("%Y-%m-%d 23:59:59").upper()
    print("future date =", toDate)
    print("future date =", toDate2)
    toDate1 = unstruct_toDate.strftime("%d-%b-%Y")    
    print("future date =", toDate3)
    fromDate4 = unstruct_fromDate.strftime("%Y-%m-%d").upper()
    
    file1="ASN_Stage_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file2="ASN_BaseT_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file3="ASN_Unmatched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file4="ASN_Matched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file5="ASN_Merged_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file6="ASN_Errors_Overdue_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file7="ASN_WMS_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file8="ASN_WMS_Unmatched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file9="ASN_WMS_Matched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file10="ASN_WMS_Merged_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"

    print(file1)
    print(file2)
    print(file3)
    print(file4)
    print(file5)   
    print(file6) 
    print(file7) 
    print(file8) 
    print(file9) 
    print(file10)  
    
    query_result1=("""
    """)
    
    query_result2=("""
          SELECT 
             Concat(mp.organization_code,msi.segment1) KEY1,
             Concat(Concat(mp.organization_code,msi.segment1),decode(rsl.source_document_code,'PO',rsh.shipment_num,decode(rsl.source_document_code,'REQ',rsh.shipment_num,ooh.order_number))) KEY2,
             Concat(Concat(Concat(mp.organization_code,msi.segment1),decode(rsl.source_document_code,'PO',rsh.shipment_num,decode(rsl.source_document_code,'REQ',rsh.shipment_num,ooh.order_number))),nvl(rsl.quantity_shipped,0)) KEY3,
             mp.organization_code,
             rsh.shipment_num,
             rsl.line_num,
             rsh.receipt_num,
             ooh.order_number,
             msi.segment1 Item,
             rsl.item_description,
             rsl.vendor_item_num,
             rsl.unit_of_measure,
             rsh.receipt_source_code,
             rsl.source_document_code,
             rsl.shipment_line_status_code,
             rsh.packing_slip,
             rsl.container_num,
             rsl.creation_date,
             trunc(rsl.creation_date) WMS_ACTIVITY_DATE_TIME,
             rsh.shipped_date,
             rsh.expected_receipt_date,
             rsl.created_by,
             rsl.shipment_header_id,
             rsl.shipment_line_id,
             nvl(rsl.quantity_received,0) quantity_received,
             nvl(rsl.quantity_shipped,0) quantity_shipped,
             rsl.po_header_id,
             rsl.po_release_id,
             rsl.po_line_id,
             rsl.po_line_location_id,
             rsl.requisition_line_id,
             rsl.req_distribution_id,
             rsl.from_organization_id,
             rsh.vendor_id,
             rsh.vendor_site_id,
             msi.inventory_item_id,
             rsh.ship_to_org_id,
             rsh.organization_id,
             rsl.to_organization_id,
             rsh.ship_to_location_id,
             rsh.ship_from_location_id,
             rsl.oe_order_header_id,
             rsl.oe_order_line_id
             ,null as RUN_TIME
             ,null as IN_WMS
             ,null as NOT_IN_WMS
             ,null as EXCEEDS_2_HR
             ,1 as TOTAL_COUNT
             ,NULL EBS_CREATION_DATE
        FROM apps.rcv_shipment_headers rsh,
             apps.rcv_shipment_lines   rsl,
             apps.mtl_system_items_b   msi,
             apps.mtl_parameters       mp,
             apps.po_vendors           pv,
             apps.oe_order_headers_all ooh
       WHERE rsl.shipment_header_id = rsh.shipment_header_id
         AND rsl.quantity_shipped = rsl.quantity_shipped
         AND msi.inventory_item_id = rsl.item_id
         AND mp.organization_id =msi.organization_id
         AND rsl.to_organization_id = msi.organization_id
         and rsl.oe_order_header_id = ooh.header_id(+)
         AND rsl.to_organization_id in (202,711,1431,1432)
         AND rsh.vendor_id = pv.vendor_id(+)
         and rsl.creation_date BETWEEN TO_DATE('"""+fromDate+"""', 'DD-MON-YYYY hh24:mi:ss')
         AND TO_DATE('"""+toDate+"""', 'DD-MON-YYYY hh24:mi:ss')
         and rsl.SHIPMENT_LINE_STATUS_CODE!='FULLY RECEIVED'
         and rsl.source_document_code in ('PO','REQ')
     """)
    
    query_result3=("""
    """)

    query_result4=("""
        select 
        concat(rh.warehouse,rd.ITEM) KEY1,
        concat(rh.warehouse,rd.ITEM,(if(rh.receipt_type='RMA',left(rd.receipt_id,10),ifnull(rh.SOURCE_FAX_NUM,'')))) KEY2,
        concat(rh.warehouse,rd.ITEM,(if(rh.receipt_type='RMA',left(rd.receipt_id,10),ifnull(rh.SOURCE_FAX_NUM,''))),round(rd.total_qty,0)) KEY3, 
        rh.warehouse, 
        rh.source_fax_num as ebs_receipt_id, 
        rh.receipt_type, 
        rh.receipt_id as scale_receipt_id,
        rd.item, 
        rh.total_lines, 
        rd.total_qty quantity_shipped, 
        rh.creation_date_time_stamp, 
        DATE_FORMAT(rh.creation_date_time_stamp, '%Y-%m-%d') WMS_ACTIVITY_DATE_TIME,
        rh.internal_receipt_num
        ,NULL IN_WMS
        ,NULL NOT_IN_WMS
        ,NULL RUN_TIME
        ,NULL EXCEEDS_2_HR
        ,1 TOTAL_COUNT
        ,null REPROCESS_COUNT
        ,NULL EBS_CREATION_DATE
        from ils_repl.RECEIPT_HEADER rh, 
             ils_repl.RECEIPT_DETAIL rd
        where rh.RECEIPT_TYPE in ('ASN','REQ')
        and rh.RECEIPT_ID=rd.RECEIPT_ID
        and rh.INTERNAL_RECEIPT_NUM=rd.INTERNAL_RECEIPT_NUM
        and rh.CLOSE_DATE is null
        and rh.CREATION_DATE_TIME_STAMP >= '"""+fromDate2+"""' 
    """)
    
    df4=pd.read_sql_query(query_result4,conn)
    header=list(df4.columns)
    print(df4.columns)
    print(file7)
    writer = pd.ExcelWriter(file7, engine='openpyxl')
    df4.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
 
    df=pd.read_sql_query(query_result2,connection)
    header=list(df.columns)
    print(df.columns)
    print(file2)
    writer = pd.ExcelWriter(file2, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    key2 = 'KEY3'
    key3 = 'KEY3'
    
    # Use pandas to read the Excel file into a DataFrame
    df2 = pd.read_excel(file7, sheet_name='Sheet1')
    df3 = pd.read_excel(file2, sheet_name='Sheet1')
    
    # Find unmatched & matched records from scale to stage
    ebs_unmatched_records = []
    ebs_matched_records = []
    
    for index, row in df3.iterrows():
        if row[key3] not in df2[key2].values:
            ebs_unmatched_record = row.to_dict()
            ebs_unmatched_record['RUN_TIME'] = pd.Timedelta(0)
            ebs_unmatched_record['IN_WMS'] = 'NOT_IN_WMS'
            ebs_unmatched_record['NOT_IN_WMS'] = 1
            ebs_unmatched_record['EXCEEDS_2_HR'] = 0  
            ebs_unmatched_records.append(ebs_unmatched_record)
        else:
            ebs_matching_row2 = df2[df2[key2] == row[key3]]
            ebs_matched_record = row.to_dict()
            ebs_matched_record['WMS_ACTIVITY_DATE_TIME'] = ebs_matching_row2['WMS_ACTIVITY_DATE_TIME'].values[0]
            ebs_matched_record['creation_date_time_stamp'] = ebs_matching_row2['creation_date_time_stamp'].values[0].astype('datetime64[ns]')
            ebs_matched_record['EBS_CREATION_DATE'] = row['CREATION_DATE']  
            ebs_matched_record['RUN_TIME'] = ebs_matching_row2['creation_date_time_stamp'].values[0].astype('datetime64[ns]')  - row['CREATION_DATE']  
            ebs_matched_record['IN_WMS'] = 'IN_WMS'
            ebs_matched_record['NOT_IN_WMS'] = 0
            ebs_matched_record['EXCEEDS_2_HR'] = int(ebs_matched_record['RUN_TIME'] > pd.Timedelta(hours=2))
            ebs_matched_records.append(ebs_matched_record)

    # Convert the lists of unmatched and matched records to DataFrames
    ebs_unmatched_records_df = pd.DataFrame(ebs_unmatched_records)
    ebs_matched_records_df = pd.DataFrame(ebs_matched_records)
   
    header=list(ebs_unmatched_records_df.columns)
    print(file8)
    writer = pd.ExcelWriter(file8, engine='openpyxl')
    ebs_unmatched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    header=list(ebs_matched_records_df.columns)
    print(file9)
    writer = pd.ExcelWriter(file9, engine='openpyxl')
    ebs_matched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    # Combine the scale_unmatched and matched records
    result_ebs_df = pd.concat([ebs_unmatched_records_df, ebs_matched_records_df], axis=0, ignore_index=True)
    result_ebs_df['WMS_ACTIVITY_DATE_TIME'] = pd.to_datetime(result_ebs_df['WMS_ACTIVITY_DATE_TIME'])
    result_ebs_df = result_ebs_df.sort_values(by=['WMS_ACTIVITY_DATE_TIME'])

    # Create a new Excel file with the combined results
    with pd.ExcelWriter(file10, engine='xlsxwriter') as writer:
        # Write the result DataFrame to a new sheet
        result_ebs_df.to_excel(writer, sheet_name='Sheet1', index=False)
    
        # Format the 'Unmatched' and 'Matched' columns
        worksheet = writer.sheets['Sheet1']
        ebs_unmatched_cond_fmt = {
            'type': 'cell',
            'criteria': 'equal to',
            'value': '"Unmatched"',
            'format': writer.book.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
        }
        ebs_matched_cond_fmt = {
            'type': 'cell',
            'criteria': 'equal to',
            'value': '"Matched"',
            'format': writer.book.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
        }
        worksheet.conditional_format('G2:G{}'.format(len(result_ebs_df)+1), ebs_unmatched_cond_fmt)
        worksheet.conditional_format('G2:G{}'.format(len(result_ebs_df)+1), ebs_matched_cond_fmt)
    
    # Find unmatched & matched records
    #unmatched_records = []
    #matched_records = []
    #for index, row in df1.iterrows():
    #    if row[key1] not in df2[key2].values:
    #        unmatched_record = row.to_dict()
    #        unmatched_record['NOT_IN_EBS'] = 1
    #        unmatched_record['EXCEEDS_2_HR'] = 0
    #        unmatched_record['IN_EBS'] = 'NOT_IN_EBS'
    #        unmatched_records.append(unmatched_record)
    #    else:
    #        #matched_record = df2[df2[key2] == row[key1]].iloc[0].to_dict()
    #        matching_row = df2[df2[key2] == row[key1]]
    #        matched_record = row.to_dict()           
    #        matched_record['CREATION_DATE'] = matching_row['CREATION_DATE'].values[0]
    #        matched_record['RUN_TIME'] = matching_row['CREATION_DATE'].values[0] - row['CREATION_DATE']
    #        matched_record['IN_EBS'] = 'IN_EBS'
    #        matched_record['NOT_IN_EBS'] = 0
    #        matched_record['EXCEEDS_2_HR'] = int(matched_record['RUN_TIME'] > pd.Timedelta(hours=2))
    #        matched_records.append(matched_record)
                
    # Convert the lists of unmatched and matched records to DataFrames
    #unmatched_records_df = pd.DataFrame(unmatched_records)
    #matched_records_df = pd.DataFrame(matched_records)
   
    #header=list(unmatched_records_df.columns)
    #print(file3)
    #writer = pd.ExcelWriter(file3, engine='openpyxl')
    #unmatched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    #writer.close()

    #header=list(matched_records_df.columns)
    #print(file4)
    #writer = pd.ExcelWriter(file4, engine='openpyxl')
    #matched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    #writer.close()
        
    # Combine the unmatched and matched records
    #result_df = pd.concat([unmatched_records_df, matched_records_df], axis=0, ignore_index=True)
    #result_df = result_df.sort_values(by=['STG_DATE'])
    
    # Create a new Excel file with the combined results
    #with pd.ExcelWriter(file5, engine='xlsxwriter') as writer:
    #    # Write the result DataFrame to a new sheet
    #    result_df.to_excel(writer, sheet_name='Combined', index=False)
    
        # Format the 'Unmatched' and 'Matched' columns
   #     worksheet = writer.sheets['Combined']
   #     unmatched_cond_fmt = {
   #         'type': 'cell',
   #         'criteria': 'equal to',
   #         'value': '"Unmatched"',
   #         'format': writer.book.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
   #     }
   #     matched_cond_fmt = {
   #         'type': 'cell',
   #         'criteria': 'equal to',
   #         'value': '"Matched"',
   #         'format': writer.book.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
   #     }
   #     worksheet.conditional_format('G2:G{}'.format(len(result_df)+1), unmatched_cond_fmt)
   #     worksheet.conditional_format('G2:G{}'.format(len(result_df)+1), matched_cond_fmt)


    # Reorder the columns in result_scale_df
    print(result_ebs_df.columns)
    column_order = ['WMS_ACTIVITY_DATE_TIME', 'IN_WMS', 'SOURCE_DOCUMENT_CODE', 'ORGANIZATION_CODE', 'TOTAL_COUNT', 'NOT_IN_WMS', 'EXCEEDS_2_HR', 'RUN_TIME', 'QUANTITY_SHIPPED']
    result_ebs_df_ordered = result_ebs_df[column_order]
    
    # Create an OrderedDict with the desired column order
    aggfunc_dict = OrderedDict([
        ('TOTAL_COUNT', 'sum'),
        ('NOT_IN_WMS', 'sum'),
        ('EXCEEDS_2_HR', 'sum'),
        ('RUN_TIME', [np.max, np.mean]),
        ('QUANTITY_SHIPPED', 'sum')
    ])
    
    # Perform the pivot operation using the ordered columns
    pt1 = pd.pivot_table(result_ebs_df_ordered.reset_index(),
                         index=['WMS_ACTIVITY_DATE_TIME', 'IN_WMS', 'SOURCE_DOCUMENT_CODE', 'ORGANIZATION_CODE'],
                         aggfunc=aggfunc_dict)
        
    # Reorder the columns based on the desired column order
    #pt1 = pt1.reindex(columns=list(aggfunc_dict.keys()))

    # Rename columns in MultiIndex
    pt1.columns = list(map('_'.join, pt1.columns.values))

    # Convert PivotTable to HTML tableORGANIZATION_CODE
    html_table4 = pt1.to_html()

    # Define the order of the columns
    #column_order = ['STG_DATE', 'IN_EBS','TRANSACTION_TYPE', 'WAREHOUSE']
       
    # Create a pivot table
    #pt = pd.pivot_table(result_df.reset_index(),
    #                index=['STG_DATE', 'IN_EBS', 'TRANSACTION_TYPE', 'WAREHOUSE'],
    #                aggfunc={'TOTAL_COUNT':'sum',
    #                         'QUANTITY':'sum',
    #                         'NOT_IN_EBS': 'sum',
    #                         'EXCEEDS_2_HR': 'sum',
    #                         'RUN_TIME': [np.max,np.mean]})

    #pt = pt.round(0)  # Round the values to 2 decimal places

    # Combine the defect percentage with the main pivot table
    #pt = pt.merge(defect_percent['DEFECT_PERCENT'], left_on='STG_DATE', right_on='STG_DATE').round(0)
    
    # Rename columns in MultiIndex
    #pt.columns = list(map('_'.join, pt.columns.values))

    # Define a function to highlight based on condition
    #def highlight_col(col):
    #    if col['NOT_IN_EBS'] > 60:
    #        return ['background-color: yellow']
    #    else:
    #        return ['background-color: white']
    
    # Define a function to highlight rows based on condition
    #def highlight_defects(row):
    #    if (row['SOA_STATUS'] == 'ERROR') or (row['SOA_STATUS'] == 'STUCK'):
    #        return ['background-color: red']
    #    else:
    #        return ['background-color: white']
        
    # Apply the style to the pivot table
    #styled_pt = pt.style.applymap(highlight_col, subset=pd.IndexSlice[:, :])
    # Apply the style to the pivot table
    #styled_pt = pt.style.applymap(highlight_defects, subset=pd.IndexSlice[:, :])
    
    # Convert PivotTable to HTML table
    #html_table2 = pt.to_html()

    # Calculate the defect percentage for each date
    # Calculate the defect percentage for each date
    result_ebs_df['WMS_ACTIVITY_DATE_TIME'] = pd.to_datetime(result_ebs_df['WMS_ACTIVITY_DATE_TIME']).dt.strftime('%Y-%m-%d')
    
    defect_percent = result_ebs_df.groupby('WMS_ACTIVITY_DATE_TIME')[['NOT_IN_WMS', 'EXCEEDS_2_HR', 'TOTAL_COUNT']].sum()
     
    # Convert the 'NOT_IN_EBS' and 'RECEIPT_COUNT' columns to numeric data types
    defect_percent['NOT_IN_WMS'] = pd.to_numeric(defect_percent['NOT_IN_WMS'], errors='coerce')
    defect_percent['EXCEEDS_2_HR'] = pd.to_numeric(defect_percent['EXCEEDS_2_HR'], errors='coerce')
    defect_percent['TOTAL_COUNT'] = pd.to_numeric(defect_percent['TOTAL_COUNT'], errors='coerce')
    
    # Drop any rows with missing or non-numeric values
    defect_percent.dropna(inplace=True)
     
    # Calculate the defect percentage for each date
    defect_percent['Errors < 1 %'] = defect_percent['NOT_IN_WMS'] / defect_percent['TOTAL_COUNT'] * 100
    defect_percent['PERCENTAGE_EXCEEDS_2_HR'] = defect_percent['EXCEEDS_2_HR'] / defect_percent['TOTAL_COUNT'] * 100
    
    if (defect_percent['EXCEEDS_2_HR'].values == 0).all():
        result1 = 0
    else:    
        result1 = (defect_percent['EXCEEDS_2_HR'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]

    if (defect_percent['NOT_IN_WMS'].values == 0).all():
        result2 = 0 
    else:
        result2 = (defect_percent['NOT_IN_WMS'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]

    result1 = float(result1)
    result1 = round(result1,2)
    result2 = float(result2)
    result2 = round(result2,2)

    defect_percent['NOT_IN_EBS'] = pd.to_numeric(defect_percent['NOT_IN_WMS'], errors='coerce') 
    
    defect_percent = defect_percent.drop(columns=['NOT_IN_WMS'])

    defect_percent = defect_percent.rename(columns={'EXCEEDS_2_HR': 'EXCEEDS_PROCESS_TIME', 'PERCENTAGE_EXCEEDS_2_HR': 'Exceeds Process Time'})
    #defect_percent = defect_percent.reset_index().rename(columns={'WMS_ACTIVITY_DATE_TIME': 'WMS_RECEIPT_DATE'})
    print(defect_percent.columns)    
     
    # Convert the defect percent data frame to an HTML table string
    html_table1 = defect_percent.to_html(index=True)
    
    
    # Open the Excel file and create an ExcelWriter object
    #writer = pd.ExcelWriter(file5, engine='openpyxl')
    #writer.book = openpyxl.load_workbook(file5)
    
    # Write the pivot table to a new sheet in the existing workbook
    #pt.to_excel(writer, sheet_name='Pivot Table')
    
    # Save the changes and close the workbook
    #writer.close()
    #writer.close()
    #import pdb

    # Set a breakpoint
    #pdb.set_trace()

    # Load the existing Excel file into a pandas DataFrame
    df_KPI = pd.read_excel('Parts_KPIs.xlsx')
    
    # Add custom text column to the pivot table DataFrame
    text_value = 'ASN'
    defect_percent.insert(0, 'Parts_KPI', text_value)
    
    # Convert pivot table rows into a DataFrame
    df_append = pd.DataFrame(defect_percent.to_records())
    
    matching = 0
    for index, row in df_KPI.iterrows():
        #fromDate4 = row['WMS_ACTIVITY_DATE_TIME']
        existing_record1 = (df_append['Parts_KPI'] == text_value) & (df_append['WMS_ACTIVITY_DATE_TIME'] == fromDate4)
        #existing_record2 = (row['Parts_KPI'] == text_value) & (datetime.strptime(row['WMS_ACTIVITY_DATE_TIME'], "%Y-%m-%d").strftime("%Y-%m-%d") == fromDate4)
        existing_record2 = (row['Parts_KPI'] == text_value) & (row['WMS_ACTIVITY_DATE_TIME'] == fromDate4)

        # Compare with index record in df_KPI
        if len(df_append[existing_record1]) > 0:
            df_KPI_record = df_KPI.loc[index]  # Get the record at the current index in df_KPI
            append_record = df_append.loc[existing_record1].iloc[0]  # Get the matching record in df_append
            
            # Compare the record values
            if df_KPI_record.equals(append_record):
                matching = 1
                # Perform any desired actions or logic for matching records
    
        # Compare with index record in df_KPI
        if existing_record2:
            #if row['Parts_KPI'] == text_value and row['WMS_ACTIVITY_DATE_TIME'].strftime("%Y-%m-%d").upper() == fromDate4:
            if row['Parts_KPI'] == text_value and row['WMS_ACTIVITY_DATE_TIME'] == fromDate4:
                # Replace existing record with df_append record
                df_KPI.loc[index] = df_append.loc[existing_record1].iloc[0]
                matching = 1
    
        # Perform other operations or comparisons as needed
    
    if matching < 1:
        # Append new record
        df_KPI = pd.concat([df_KPI, df_append], ignore_index=False)
        
    
    # Convert 'Exceeds Process Time' column to numeric
    #df_KPI['Exceeds Process Time'] = pd.to_numeric(df_KPI['Exceeds Process Time'], errors='coerce')
    
    # Apply formatting to 'Exceeds Process Time' column
    #df_KPI['Exceeds Process Time'] = df_KPI['Exceeds Process Time'].apply(lambda x: "{:.1f}%".format(x) if pd.notnull(x) else "")        
    
    # Save the updated DataFrame to a new Excel file
    df_KPI.to_excel('Parts_KPIs.xlsx', index=False)
       
    file0=file5
    usecols= []
    team = 'OPS Team'
    subject = "Parts_KPI_ASN_{}"+toDate1
    emailbodytext = 'Please find attached Details for the Parts KPI ASN,  Request you to Report & Clear the  Defects Which are highlighted.'
    files = [file2,file7,file8,file9,file10]
    
    email_proc(sender_email, receiv_email, cc_email, passwd_email, unstruct_toDate, file0, usecols, team, attention, emailbodytext, files, subject, html_table1, html_table4, result1, result2)


    # Construct the file pattern to match
    file_pattern = 'ASN*.xlsx'
    
    # Get a list of file paths that match the pattern
    file_paths = glob.glob(os.path.join(working_folder, file_pattern))
    
    # Iterate over the file paths and delete each file
    for file_path in file_paths:
        os.remove(file_path)


def email_proc(sender_email, receiv_email, cc_email, passwd_email, unstruct_toDate, file0, usecols, team, attention, emailbodytext, files, subject, html_table1, html_table4, result1, result2):

    import pandas as pd
    from datetime import datetime, timedelta
    from email.message import EmailMessage
    import smtplib
    import os.path
    import ssl
    #from passwd import pswd
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from openpyxl import load_workbook
    import openpyxl
    
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiv_email
    msg['CC'] = cc_email
    msg['Subject'] = subject
    
    # Read the CSV file into a pandas DataFrame
    #df = pd.read_csv(file0)
    
    # Convert the pandas DataFrame to a CSV file
    #df.to_csv(file0, index=False)
    
    # Read CSV file and extract desired rows and columns
    #df = pd.read_csv(file0, usecols=usecols)
    #nrows = df.shape[0]
    
    # Read CSV file again using nrows
    #df = pd.read_csv(file0, usecols=usecols, nrows=nrows)

    # Load PivotTable from Excel file
    #workbook = openpyxl.load_workbook(file0)
    #worksheet = workbook['Pivot Table']
    #pivot_table = worksheet['A1'].pivotTableDefinition.cacheDefinition
    
    # Convert PivotTable to pandas DataFrame
    #df = pd.read_excel(file0, sheet_name='Pivot Table', index_col=0, skiprows=pivot_table.row_cache_start, nrows=pivot_table.row_count)
    
    # Generate HTML table from DataFrame
    #html_table = df.to_html()
    
    # Convert DataFrame to HTML table
    #html_table = df.to_html(index=False)
    
    # Convert DataFrame to styled HTML table
    #html_table = (df.style
    #    .set_properties(**{'border-collapse': 'collapse', 'border': '1px solid black'})
    #    .set_table_styles([{'selector': 'th', 'props': [('background-color', 'lightblue'),
    #                                                     ('font-weight', 'bold'),
    #                                                     ('border', '1px solid black')]},
    #                       {'selector': 'tr:last-child', 'props': [('background-color', 'lightblue'),
    #                                                                ('font-weight', 'bold'),
    #                                                                ('border', '1px solid black')]}])
    #    .set_table_attributes('border="1" class="dataframe table table-striped table-hover table-sm" style="text-align: center; font-family: Arial"')
    #    .set_table_attributes([{'selector': 'thead tr',
    #                            'props': [('background-color', 'lightblue')]}])
    #    .set_table_attributes([{'selector': 'tbody tr:nth-child(1), tbody tr:nth-child(2)',
    #                            'props': [('background-color', 'lightblue')]}])
    #    .render()
    #)
    
    # Add HTML message to email body
    html_content = f"""
    <html>
        <body>
            <p> </p>
            <p> Kind attention required from: {attention}</p>
            <p> </p>
            <p>KPI For ASN<p>
            <p>  1. ASN should get processed within 2 hours (EBS TO WMS) :-  <b>{result1} % ASN are Processed after 2 hours</b> <p>
            <p>  2. Errors should be Less than 1% error                                      :-  <b>{result2} % Errors</b> <p>
            <p> </p>
            <p>{emailbodytext}</p>
            <p> </p>
            <p>Parts Automation-Python</p>
            <p> </p>
        </body>
    </html>
    """
    
    # Add HTML table to email body
    html_body = MIMEText(html_content + html_table1 , 'html')
    msg.attach(html_body)
    
    # Attach multiple files to email
    files = files
    
    # List of paths to files
    for file in files:
        with open(file, 'rb') as f:
            file_content = f.read()
            # Create attachment and add to message
            attachment = MIMEApplication(file_content, Name=file)
            attachment['Content-Disposition'] = f'attachment; filename="{file}"'
            msg.attach(attachment)
    
    # Send email
    #with smtplib.SMTP('smtp.office365.com', 587) as smtp:
    #    context = ssl.create_default_context()
    #    smtp.ehlo()
    #    smtp.starttls(context=context)
    #    smtp.login(sender_email, passwd_email)
    #    smtp.sendmail(sender_email, receiv_email.split(','), msg.as_string())
    with smtplib.SMTP('smtp.office365.com', 587) as smtp:
        context = ssl.create_default_context()
        smtp.ehlo()
        smtp.starttls(context=context)
        smtp.login(sender_email, passwd_email)
        receiv_email_list = receiv_email.split(',')
        cc_email_list = cc_email.split(',')
        receiv_email_list.extend(cc_email_list)
        smtp.sendmail(sender_email, receiv_email_list, msg.as_string())        

        print("job has been succesfully executed ", datetime.now())

   
def main():
    import sys
    import datetime
    sub_proc1()
    
if __name__ == "__main__":
    main()