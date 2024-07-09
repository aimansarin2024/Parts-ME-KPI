def sub_proc1():
    import pandas as pd
    import cx_Oracle as ora
    import pymysql
    import sys ; sys.setrecursionlimit(sys.getrecursionlimit() * 5)
    import csv
    import os
    import glob
    from datetime import datetime, timedelta
    import pandas as pd
    import openpyxl  
    import numpy as np
    import configparser
    import openpyxl
    from cryptography.fernet import Fernet
    import configparser
    from collections import OrderedDict
    import time
    os.chdir("C:\Parts POSB Automation")
    # Load the encryption key
    with open('encryption_key.key', 'rb') as file:
        key = file.read()
    
    # Create a Fernet cipher using the key
    cipher_suite = Fernet(key)
    
    # Read the encrypted data from the encrypted_config.ini file
    with open('encrypted_config.ini', 'rb') as file:
        ciphertext = file.read()
    
    # Decrypt the ciphertext using the cipher
    plaintext = cipher_suite.decrypt(ciphertext)
    
    # Convert the plaintext to a string
    plaintext_str = plaintext.decode('utf-8')
    
    # Parse the decrypted config data into a dictionary
    config_data = configparser.ConfigParser(dict_type=OrderedDict)
    config_data.read_string(plaintext_str)
    
    credentials = config_data['Credentials']
    connstr_plsql_v = credentials['connstr_plsql']
    connstr_mysql_v = credentials['connstr_mysql']
    host_mysql_v = credentials['host_mysql']
    user_mysql_v = credentials['user_mysql']
    pswd_mysql_v = credentials['pswd_mysql']
    dbas_mysql_v = credentials['dbas_mysql']
    passwd_email = credentials['passwd_email']
    sender_email = credentials['sender_email']
    receiv_email = credentials['receiv_email']
    cc_email = credentials['cc_email']
    attention = credentials['attention0']
    working_folder = credentials['working_folder']
    os.chdir(working_folder)

    print(working_folder)

    ora.init_oracle_client(
        lib_dir=r"C:\cx_Oracle\instantclient-basic-windows.x64-21.8.0.0.0dbru\instantclient-basic-windows.x64-21.8.0.0.0dbru\instantclient_21_8")
    #connstr = 'RO999999999/xxxxxxxx@geaebsdbadg1.appl.ge.com:1521/ercebs1p'
    connpl = ora.connect(connstr_plsql_v)
 
    connmy=pymysql.connect(host=host_mysql_v, user=user_mysql_v, passwd=pswd_mysql_v, database=dbas_mysql_v)
    
    now = datetime.now()
    # Check if today is Monday
    if now.weekday() == 0:
        # If it is Monday, set unstruct_fromDate to two days ago
        unstruct_fromDate = now + timedelta(days=-3)
        unstruct_fromDate3 = now + timedelta(days=-10)                              
        fromDate3 = unstruct_fromDate3.strftime("%d-%b-%Y 00:00:01").upper()
        #print("date and time =", fromDate3)
        unstruct_toDate3 = now + timedelta(days=-4)                              
        toDate3 = unstruct_toDate3.strftime("%d-%b-%Y 23:59:59").upper()
    else:
        # If it is not Monday, set unstruct_fromDate to one day ago
        unstruct_fromDate = now + timedelta(days=-1)
        unstruct_fromDate3 = now + timedelta(days=-8)                              
        fromDate3 = unstruct_fromDate3.strftime("%d-%b-%Y 00:00:01").upper()
        #print("date and time =", fromDate3)
        unstruct_toDate3 = now + timedelta(days=-2)                              
        toDate3 = unstruct_toDate3.strftime("%d-%b-%Y 23:59:59").upper()

    unstruct_toDate = now + timedelta(days=-1)           
    fromDate = unstruct_fromDate.strftime("%d-%b-%Y 00:00:01").upper()
    fromDate2 = unstruct_fromDate.strftime("%Y-%m-%d 00:00:01").upper()
    print("From Date =", fromDate)
    #print("date and time =", fromDate2)
    toDate = unstruct_toDate.strftime("%d-%b-%Y 23:59:59").upper()
    toDate2 = unstruct_toDate.strftime("%Y-%m-%d 23:59:59").upper()
    print("To Date =", toDate)
    #print("future date =", toDate2)
    toDate1 = unstruct_toDate.strftime("%d-%b-%Y")    
    #print("future date =", toDate3)
    fromDate4 = unstruct_fromDate.strftime("%Y-%m-%d 00:00:00").upper()

      
    file11="EBS_SO_download_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file12="SCALE_SO_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file13="SO_Missing_In_WMS_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file14="SO_in_WMS_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file15="SO_Combine_Data_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file16="IRMS_SO_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    #file23="SHIP_CFM_Missing_In_EBS_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".csv"
    #file24="SHIP_CFM_Missing_in_IRMS_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".csv"
   
    
    # IRMS DATA EXTRACT
    
    
#    query_result2=("""
#    select concat(concat(concat(a.warehouse,a.item),a.erp_order_number),a.quantity) matchkey, a.* 
#    from irms_data.ship_confirmations_cons a
#    where a.date_and_time between '"""+fromDate1+""""' and '"""+toDate1+""""' and a.item !='0'
#    """)
                  
#    df=pd.read_sql_query(query_result1,connmy)
#    header=list(df.columns)
#    print(file11)
#    df.to_csv(file11,index=False,header=header)
    
#    df=pd.read_sql_query(query_result2,connmy)
#    header=list(df.columns)
#    print(file21)
#    df.to_csv(file21,index=False,header=header)
    
    # EBS DATA EXTRACT
    
    query_result1=("""
        SELECT
            CONCAT(MTP.ORGANIZATION_CODE, msi.segment1) AS KEY1,
            CONCAT(CONCAT(MTP.ORGANIZATION_CODE, msi.segment1), TO_CHAR(OH.ORDER_NUMBER)) AS KEY2,
            CONCAT(CONCAT(CONCAT(MTP.ORGANIZATION_CODE, msi.segment1), TO_CHAR(OH.ORDER_NUMBER)), ABS(OL.ORDERED_QUANTITY)) AS KEY3,
            CONCAT(MTP.ORGANIZATION_CODE, TO_CHAR(OH.ORDER_NUMBER)) AS KEY4,
            MTP.ORGANIZATION_CODE,
            OH.ORDER_NUMBER ORDER_NUMBER,
            OL.ORDERED_ITEM,
            msi.segment1 Inventory_item,
            OL.ORDERED_QUANTITY,
            OL.UNIT_SELLING_PRICE,
            OL.ORDERED_QUANTITY * OL.UNIT_SELLING_PRICE AS ORDER_VALUE,
            OL.SCHEDULE_SHIP_DATE,
            CASE
                WHEN hstg.creation_date BETWEEN (TRUNC(hstg.creation_date) + INTERVAL '0 00:00:01' DAY TO SECOND)
                     AND (TRUNC(hstg.creation_date) + INTERVAL '0 11:00:00' DAY TO SECOND) THEN 'Before 11am EST'
                WHEN hstg.creation_date BETWEEN (TRUNC(hstg.creation_date) + INTERVAL '0 11:00:01' DAY TO SECOND)
                     AND (TRUNC(hstg.creation_date) + INTERVAL '0 14:00:00' DAY TO SECOND) THEN 'Before 2pm EST'
                ELSE 'EST 2pm After'
            END AS Shipment_Lot,
            ROUND(((soa_to_wms_timestamp - hstg.creation_date) * 24 * 60), 0) AS MAX_RUN_TIME_MIN,
            ROUND(((soa_to_wms_timestamp - hstg.creation_date) * 24 * 60), 0) AS AVG_RUN_TIME_MIN,
            ROUND(((hstg.wms_to_soa_timestamp - hstg.soa_to_wms_timestamp) * 24 * 60), 0) AS Download_Minutes,
            CASE
                WHEN hstg.creation_date >= (TRUNC(hstg.creation_date) + INTERVAL '0 00:00:01' DAY TO SECOND)
                     AND hstg.creation_date <= (TRUNC(hstg.creation_date) + INTERVAL '0 10:40:00' DAY TO SECOND)
                     AND hstg.soa_to_wms_timestamp > (TRUNC(hstg.creation_date) + INTERVAL '0 11:00:00' DAY TO SECOND) THEN 1
                WHEN hstg.creation_date >= (TRUNC(hstg.creation_date) + INTERVAL '0 10:40:01' DAY TO SECOND)
                     AND hstg.creation_date <= (TRUNC(hstg.creation_date) + INTERVAL '0 13:40:00' DAY TO SECOND)
                     AND hstg.soa_to_wms_timestamp > (TRUNC(hstg.creation_date) + INTERVAL '0 14:00:00' DAY TO SECOND) THEN 1
                ELSE 0
            END AS AFTER_CUTOFF,
            --hstg.creation_date AS STG_RELEASED_DATE,
            ol.creation_date AS ol_creation_date,
            TRUNC(hstg.creation_date) AS CREATION_DATET,
            hstg.creation_date AS stg_creation_date,
            ol.last_update_date AS ol_last_update_Date,
            wdd.creation_date AS WDD_CREATION_DATE,
            wdd.last_update_date AS wdd_last_update_date,
            hstg.soa_to_wms_timestamp,
            hstg.wms_to_soa_timestamp,
            OH.SHIPMENT_PRIORITY_CODE,
            OL.SHIPPING_METHOD_CODE,
            OL.DEMAND_CLASS_CODE,
            OL.FREIGHT_TERMS_CODE,
            OL.SCHEDULE_STATUS_CODE,
            OL.SOURCE_TYPE_CODE,
            OL.CANCELLED_FLAG,
            OL.BOOKED_FLAG,
            OL.OPEN_FLAG,
            OL.SHIPPING_INTERFACED_FLAG,
            COALESCE(OL.FLOW_STATUS_CODE, 'Not_Selected') AS FLOW_STATUS_CODE,
            hstg.release_number,
            hstg.status AS stag_status,
            hstg.soa_status,
            hstg.soa_error_message,
            hstg.process_message,
            OH.ATTRIBUTE15,
            OH.HEADER_ID,
            OL.LINE_ID,
            NULL AS WMS_ACTIVITY_DATE_TIME,
            NULL AS RUN_TIME,
            NULL AS EXCEEDS_1_HR,
            NULL AS IN_WMS,
            NULL AS NOT_IN_WMS,
            WDD.RELEASED_STATUS,
            WDD.DELIVERY_DETAIL_ID,
            1 AS TOTAL_COUNT
        FROM
            APPS.GEWSH_APL_SO_OUT_STG_TBL hstg
            JOIN APPS.GEWSH_APL_SO_LNS_STG_TBL LSTG ON HSTG.BATCH_ID = LSTG.BATCH_ID
            JOIN APPS.OE_ORDER_HEADERS_ALL OH ON LSTG.HEADER_ID = OH.HEADER_ID
            JOIN APPS.OE_ORDER_LINES_ALL OL ON LSTG.LINE_ID = OL.LINE_ID
            JOIN APPS.WSH_DELIVERY_DETAILS WDD ON OL.LINE_ID = WDD.SOURCE_LINE_ID
            JOIN APPS.MTL_PARAMETERS MTP ON OL.SHIP_FROM_ORG_ID = MTP.ORGANIZATION_ID
            JOIN APPS.mtl_system_items_b MSI ON mtp.organization_id =msi.organization_id and ol.inventory_item_id = msi.inventory_item_id
        WHERE 1=1
            and hstg.creation_date between TO_DATE('"""+fromDate+"""', 'DD-MON-YYYY hh24:mi:ss') and TO_DATE('"""+toDate+"""', 'DD-MON-YYYY hh24:mi:ss')
            AND OH.ATTRIBUTE15 IS NOT NULL
            AND MTP.ORGANIZATION_CODE IN ('JEF', 'RVR', 'WAL', 'NAP')
            AND WDD.RELEASED_STATUS = 'Y'
            AND OH.HEADER_ID = WDD.SOURCE_HEADER_ID    
    """)
    
    query_result2=("""
        SELECT 
        concat(oh.warehouse,od.item) KEY1,
        concat(concat(oh.warehouse,od.item),SUBSTRING_INDEX(oh.ERP_ORDER,'-',1)) KEY2,
        concat(concat(concat(oh.warehouse,od.item),SUBSTRING_INDEX(oh.ERP_ORDER,'-',1)),ROUND(abs(od.open_QTY),0)) KEY3,
        concat(oh.warehouse,SUBSTRING_INDEX(oh.ERP_ORDER,'-',1)) KEY4,
        oh.warehouse, oh.ERP_order, od.user_def5, oh.internal_order_num, oh.condition, oh.order_type, oh.total_shipments,oh.containers_Shipped,
        oh.creation_Date_time_stamp as WMS_ACTIVITY_DATE_TIME,oh.date_time_stamp, oh.customer, oh.customer_name,oh.ship_to_Country, oh.customer_category7, od.item,
        od.open_qty, od.item_net_price
        FROM ils_repl.ORDER_HEADER oh join ils_repl.ORDER_DETAIL od
        on oh.warehouse = od.warehouse and oh.internal_order_num = od.internal_order_num
        WHERE oh.ERP_ORDER like '%EBS'
        AND OD.USER_DEF5 REGEXP '[0-9]+'
        AND DATE_FORMAT(CREATION_DATE_TIME_STAMP,'%Y-%m-%d %h:%i:%s') >= '"""+fromDate2+"""' 
    """)             
      
    query_result3=("""
        select 
        concat(concat(warehouse,SUBSTRING_INDEX(erp_order_number,'-',1))) KEY4,
        warehouse, erp_order_number, date_and_time WMS_ACTIVITY_DATE_TIME
           from irms_data.orders
        where (erp_order_number not like '%-1-%-0') -- excluding COPS RMAs
         and (erp_order_number like '%EBS') -- inluding EBS RMAs, assuming ASNs/IRNs are  created ending with EBS
         and DATE_FORMAT(date_and_time,'%Y-%m-%d %h:%i:%s') >= '"""+fromDate2+"""' 
    """)          

    # Export to Excel
    df=pd.read_sql_query(query_result1,connpl)
    header=list(df.columns)
    print(file11)
    writer = pd.ExcelWriter(file11, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()

    # Export to Excel
    df=pd.read_sql_query(query_result2,connmy)
    header=list(df.columns)
    print(file12)
    writer = pd.ExcelWriter(file12, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    # Export to Excel
    df=pd.read_sql_query(query_result3,connmy)
    header=list(df.columns)
    print(file16)
    writer = pd.ExcelWriter(file16, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
#*************************************
    
    
    key1 = 'KEY2'
    key2 = 'KEY2'
    key3 = 'KEY4'   
    key4 = 'KEY4'   
    
    # Use pandas to read the Excel file into a DataFrame
    df1 = pd.read_excel(file11, sheet_name='Sheet1')
    df2 = pd.read_excel(file12, sheet_name='Sheet1')
    df3 = pd.read_excel(file16, sheet_name='Sheet1')
    
    # Find unmatched & matched records from scale to stage
    scale_unmatched_records = []
    scale_matched_records = []
    
    for index, row in df1.iterrows():
        if row[key1] not in df2[key2].values and row[key3] not in df3[key4].values:
            scale_unmatched_record = row.to_dict()
            scale_unmatched_record['RUN_TIME'] = pd.Timedelta(0)
            scale_unmatched_record['IN_WMS'] = 'NOT_IN_WMS'
            scale_unmatched_record['NOT_IN_WMS'] = 1
            scale_unmatched_record['EXCEEDS_1_HR'] = 0  
            scale_unmatched_records.append(scale_unmatched_record)
            
        elif row[key1] in df2[key2].values and row[key3] not in df3[key4].values:
            scale_matching_row1 = df2[df2[key2] == row[key1]]
            scale_matched_record = row.to_dict()
            scale_matched_record['WMS_ACTIVITY_DATE_TIME'] = scale_matching_row1['WMS_ACTIVITY_DATE_TIME'].values[0]
            scale_matched_record['RUN_TIME'] = scale_matching_row1['WMS_ACTIVITY_DATE_TIME'].values[0].astype('datetime64[ns]') - row['STG_CREATION_DATE']
            scale_matched_record['IN_WMS'] = 'IN_EBS'
            scale_matched_record['NOT_IN_WMS'] = 0
            scale_matched_record['EXCEEDS_1_HR'] = int(scale_matched_record['RUN_TIME'] > pd.Timedelta(hours=1))
            
            if row['STG_CREATION_DATE'] >= (row['STG_CREATION_DATE'].replace(hour=0, minute=0, second=1)) \
                    and row['STG_CREATION_DATE'] <= (row['STG_CREATION_DATE'].replace(hour=10, minute=40, second=0)) \
                    and scale_matched_record['WMS_ACTIVITY_DATE_TIME'] > (row['STG_CREATION_DATE'].replace(hour=11, minute=0, second=0)):
                scale_matched_record['AFTER_CUTOFF'] = 1
            elif row['STG_CREATION_DATE'] >= (row['STG_CREATION_DATE'].replace(hour=10, minute=40, second=1)) \
                    and row['STG_CREATION_DATE'] <= (row['STG_CREATION_DATE'].replace(hour=13, minute=40, second=0)) \
                    and scale_matched_record['WMS_ACTIVITY_DATE_TIME'] > (row['STG_CREATION_DATE'].replace(hour=14, minute=0, second=0)):
                scale_matched_record['AFTER_CUTOFF'] = 1
            else:
                scale_matched_record['AFTER_CUTOFF'] = 0
            
            scale_matched_records.append(scale_matched_record)    
            
        elif row[key1] not in df2[key2].values and row[key3] in df3[key4].values:
            scale_matching_row1 = df3[df3[key4] == row[key3]]
            scale_matched_record = row.to_dict()
            scale_matched_record['WMS_ACTIVITY_DATE_TIME'] = scale_matching_row1['WMS_ACTIVITY_DATE_TIME'].values[0]
            scale_matched_record['RUN_TIME'] = scale_matching_row1['WMS_ACTIVITY_DATE_TIME'].values[0].astype('datetime64[ns]') - row['STG_CREATION_DATE']
            scale_matched_record['IN_WMS'] = 'IN_EBS'
            scale_matched_record['NOT_IN_WMS'] = 0
            scale_matched_record['EXCEEDS_1_HR'] = int(scale_matched_record['RUN_TIME'] > pd.Timedelta(hours=1))
            
            if row['STG_CREATION_DATE'] >= (row['STG_CREATION_DATE'].replace(hour=0, minute=0, second=1)) \
                    and row['STG_CREATION_DATE'] <= (row['STG_CREATION_DATE'].replace(hour=10, minute=40, second=0)) \
                    and scale_matched_record['WMS_ACTIVITY_DATE_TIME'] > (row['STG_CREATION_DATE'].replace(hour=11, minute=0, second=0)):
                scale_matched_record['AFTER_CUTOFF'] = 1
            elif row['STG_CREATION_DATE'] >= (row['STG_CREATION_DATE'].replace(hour=10, minute=40, second=1)) \
                    and row['STG_CREATION_DATE'] <= (row['STG_CREATION_DATE'].replace(hour=13, minute=40, second=0)) \
                    and scale_matched_record['WMS_ACTIVITY_DATE_TIME'] > (row['STG_CREATION_DATE'].replace(hour=14, minute=0, second=0)):
                scale_matched_record['AFTER_CUTOFF'] = 1
            else:
                scale_matched_record['AFTER_CUTOFF'] = 0

            scale_matched_records.append(scale_matched_record)


    # Convert the lists of unmatched and matched records to DataFrames
    scale_unmatched_records_df = pd.DataFrame(scale_unmatched_records)
    scale_matched_records_df = pd.DataFrame(scale_matched_records)
   
    header=list(scale_unmatched_records_df.columns)
    print(file13)
    writer = pd.ExcelWriter(file13, engine='openpyxl')
    scale_unmatched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    header=list(scale_matched_records_df.columns)
    print(file14)
    writer = pd.ExcelWriter(file14, engine='openpyxl')
    scale_matched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    # Combine the scale_unmatched and matched records
    result_scale_df = pd.concat([scale_unmatched_records_df, scale_matched_records_df], axis=0, ignore_index=True)
    #result_scale_df = result_scale_df.sort_values(by=['WMS_ACTIVITY_DATE_TIME'])

    # Create a new Excel file with the combined results
    with pd.ExcelWriter(file15, engine='xlsxwriter') as writer:
        # Write the result DataFrame to a new sheet
        result_scale_df.to_excel(writer, sheet_name='Sheet1', index=False)
    

#*************************************

    # Combine the unmatched and matched records
    result_df =  result_scale_df


    # Load the Excel file into a pandas dataframe
    #xl = pd.ExcelFile(file11)
    #df = xl.parse(xl.sheet_names[0]) # Read the first sheet, change the sheet name if required
    
    #print(df.columns)
    # Define the order of the columns
    column_order = ['ORDER_NUMBER', 'IN_WMS', 'NOT_IN_WMS', 'RUN_TIME','AFTER_CUTOFF', 'MAX_RUN_TIME_MIN', 'AVG_RUN_TIME_MIN', 'ORDERED_QUANTITY','ORDER_VALUE']

    # Create a pivot table
    #pt = pd.pivot_table(df, values=['ORDER_NUMBER','AFTER_CUTOFF','PICKED_DAYS','PICKED_MINUTES', 'INTERFACED_MINUTES'], index=['WPB_CREATION_DATET','SHIPMENT_LOT','FLOW_STATUS_CODE','SOA_STATUS','STAG_STATUS'],aggfunc={'ORDER_NUMBER':'count','AFTER_CUTOFF':'sum','PICKED_MINUTES': [np.mean,np.max,np.min],'PICKED_DAYS': [np.mean,np.max,np.min],'PICKED_MINUTES': [np.mean,np.max,np.min],'INTERFACED_MINUTES': [np.mean,np.max,np.min]})
    # Create a pivot table
    pt = pd.pivot_table(result_df, 
                    values=column_order,     
                    index=['CREATION_DATET','ORGANIZATION_CODE','SHIPMENT_LOT','FLOW_STATUS_CODE','RELEASED_STATUS','SOA_STATUS','STAG_STATUS','SOA_ERROR_MESSAGE'],
                    aggfunc={'ORDER_NUMBER':'count',
                             'NOT_IN_WMS': 'sum',
                             #'EXCEEDS_1_HR':'sum',
                             'AFTER_CUTOFF':'sum',
                             #'RUN_TIME':[np.max, np.mean],
                             #'MAX_RUN_TIME_MIN': [np.max],
                             #'AVG_RUN_TIME_MIN': [np.mean],                             
                             'ORDERED_QUANTITY':'sum',
                             'ORDER_VALUE':'sum'})
                             #'C_PICKED_MINUTES': [np.mean,np.max,np.min],
                             #'D_PICKED_DAYS': [np.mean,np.max,np.min],
                             #'E_INTERFACED_MINUTES': [np.mean,np.max,np.min]})
       
   
    pt = pt.round(0)  # Round the values to 2 decimal places
    # Round the mean values to 0 decimal places
    pt = pt.round({'C_MAX_RUN_TIME_MIN': 0, 'D_AVG_RUN_TIME_MIN': 0})
    
    # Rename columns in MultiIndex
    pt.columns = list(map('_'.join, pt.columns.values))
    

    
    # Define a function to highlight based on condition
    def highlight_col(col):
        if col['AFTER_CUTOFF'] > 60:
            return ['background-color: yellow']
        else:
            return ['background-color: white']
    
    # Define a function to highlight rows based on condition
    def highlight_defects(row):
        if (row['SOA_STATUS'] == 'ERROR') or (row['SOA_STATUS'] == 'STUCK'):
            return ['background-color: red']
        else:
            return ['background-color: white']
        
    # Apply the style to the pivot table
    #styled_pt = pt.style.applymap(highlight_col, subset=pd.IndexSlice[:, :])
    # Apply the style to the pivot table
    #styled_pt = pt.style.applymap(highlight_defects, subset=pd.IndexSlice[:, :])

    
    # Convert PivotTable to HTML table
    html_table2 = pt.to_html()
    #print(pt.columns)
    
    # Export pivot table to the same Excel file in a new sheet
    #writer = pd.ExcelWriter(file11, engine='openpyxl')
    # Load the existing workbook
    # Load the existing workbook
    # Load the existing workbook
    #existing_workbook = openpyxl.load_workbook(file11)
    
    # Create a new workbook and copy sheets from the existing workbook
    #new_workbook = openpyxl.Workbook()
    #for sheet in existing_workbook.sheetnames:
    #    existing_sheet = existing_workbook[sheet]
    #    new_sheet = new_workbook.create_sheet(title=sheet)
    #    for row in existing_sheet.iter_rows():
    #        for cell in row:
    #            new_sheet[cell.coordinate].value = cell.value
    
    # Write the pivot table to the appropriate sheet in the new workbook
    #pt.to_excel(new_sheet, sheet_name='Pivot Table')
    
    # Save the modified workbook
    #new_workbook.save(filename=file11)

    # Calculate the defect percentage for each date
    defect_percent = result_df.groupby('CREATION_DATET')[['AFTER_CUTOFF', 'NOT_IN_WMS', 'TOTAL_COUNT']].sum()
   
    # Convert the 'B_AFTER_CUTOFF' and 'AFTER_CUTOFF' columns to numeric data types
    defect_percent['AFTER_CUTOFF'] = pd.to_numeric(defect_percent['AFTER_CUTOFF'], errors='coerce')
    defect_percent['NOT_IN_WMS'] = pd.to_numeric(defect_percent['NOT_IN_WMS'], errors='coerce')
    defect_percent['TOTAL_COUNT'] = pd.to_numeric(defect_percent['TOTAL_COUNT'], errors='coerce')
   
    # Drop any rows with missing or non-numeric values
    defect_percent.dropna(inplace=True)

    print()
    # Calculate the defect percentage for each date
    defect_percent['PERCENTAGE_AFTER_CUTOFF'] = defect_percent['AFTER_CUTOFF'] / defect_percent['TOTAL_COUNT'] * 100
    defect_percent['Errors < 1 %'] = defect_percent['NOT_IN_WMS'] / defect_percent['TOTAL_COUNT'] * 100
    result1 = (defect_percent['AFTER_CUTOFF'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]
    result2 = (defect_percent['NOT_IN_WMS'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]
    
    result1 = float(result1)
    result1 = round(result1,2)
    
    result2 = float(result2)
    result2 = round(result2,2)
    
    # Convert the defect percent data frame to an HTML table string 
    html_table1 = defect_percent.to_html(index=True)
    
    defect_percent = defect_percent.rename(columns={'NOT_IN_WMS': 'NOT_IN_EBS','AFTER_CUTOFF': 'EXCEEDS_PROCESS_TIME', 'PERCENTAGE_AFTER_CUTOFF': 'Exceeds Process Time'})
    defect_percent = defect_percent.reset_index().rename(columns={'CREATION_DATET': 'STAGE_DATE'})
    print(defect_percent.columns)   

            
    # Open the Excel file and create an ExcelWriter object
    #writer = pd.ExcelWriter(file5, engine='openpyxl')
    #writer.book = openpyxl.load_workbook(file5)
   
    # Write the pivot table to a new sheet in the existing workbook
    #pt.to_excel(writer, sheet_name='Pivot Table')
   
    # Save the changes and close the workbook
    #writer.close()
    #writer.close()

    # Rename columns in pivot table
    #defect_percent = defect_percent.rename(columns={'WPB_LAST_UPDATE_DATET': 'CREATION_DATE', 'AFTER_CUTOFF': 'AFTER_CUTOFF', 'TOTAL_COUNT': 'TOTAL_COUNT', 'DEFECT_PERCENT': 'DEFECT_PERCENT'})

    # Load the existing Excel file into a pandas DataFrame
    df_KPI = pd.read_excel('Parts_KPIs.xlsx')
    
    # Add custom text column to the pivot table DataFrame
    text_value = 'Order Downloads'
    defect_percent.insert(0, 'Parts_KPI', text_value)
    
    # Convert pivot table rows into a DataFrame

    df_append = pd.DataFrame(defect_percent.to_records())
    df_append = df_append.reset_index().rename(columns={'STAGE_DATE': 'WMS_ACTIVITY_DATE_TIME'})
    df_append['WMS_ACTIVITY_DATE_TIME'] = df_append['WMS_ACTIVITY_DATE_TIME'].apply(lambda row: datetime.strptime(str(row), "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d"))
    print(df_append.columns)    

    # Iterate over rows in df_append
    matching = 0
    for index, row in df_KPI.iterrows():
        #fromDate4 = row['WMS_ACTIVITY_DATE_TIME']

        existing_record1 = (df_append['Parts_KPI'] == text_value) & (df_append['WMS_ACTIVITY_DATE_TIME'] == fromDate4)
        #existing_record2 = (row['Parts_KPI'] == text_value) & (datetime.strptime(row['WMS_ACTIVITY_DATE_TIME'], "%Y-%m-%d").strftime("%Y-%m-%d") == fromDate4)
        existing_record2 = (row['Parts_KPI'] == text_value) & (row['WMS_ACTIVITY_DATE_TIME'] == fromDate4)
        
             
        # Compare with index record in df_KPI
        if len(df_append[existing_record1]) > 0:
            df_KPI_record = df_KPI.loc[index]  # Get the record at the current index in df_KPI
            append_record = df_append.loc[existing_record1].iloc[0]  # Get the matching record in df_append
            
            # Compare the record values
            if df_KPI_record.equals(append_record):
                matching = 1
                # Perform any desired actions or logic for matching records
    
        # Compare with index record in df_KPI
        if existing_record2:
            #if row['Parts_KPI'] == text_value and row['WMS_ACTIVITY_DATE_TIME'].strftime("%Y-%m-%d").upper() == fromDate4:
            if row['Parts_KPI'] == text_value and row['WMS_ACTIVITY_DATE_TIME'] == fromDate4:
                # Replace existing record with df_append record
                df_KPI.loc[index] = df_append.loc[existing_record1].iloc[0]
                matching = 1
    
        # Perform other operations or comparisons as needed
    
    if matching < 1:
        # Append new record
        df_KPI = pd.concat([df_KPI, df_append], ignore_index=False)
            
    # Convert 'Exceeds Process Time' column to numeric
    #df_KPI['Exceeds Process Time'] = pd.to_numeric(df_KPI['Exceeds Process Time'], errors='coerce')
    
    # Apply formatting to 'Exceeds Process Time' column
    #df_KPI['Exceeds Process Time'] = df_KPI['Exceeds Process Time'].apply(lambda x: "{:.1f}%".format(x) if pd.notnull(x) else "")        
    
    # Save the updated DataFrame to a new Excel file
    df_KPI.to_excel('Parts_KPIs.xlsx', index=False)

    file0=file11
    usecols= []
    team = 'OPS Team'
    subject = "Parts_KPI_Order_Download_{}"+toDate1
    emailbodytext = 'Please find attached Details for the Parts KPI Sales Order DownLoad,  Request you to Report & Clear the  Defects Which are highlighted.'
    #files = [file11,file12,file13,file14,file15,file16]
    files = [file13]

    email_proc(sender_email, receiv_email, cc_email, passwd_email, 
               unstruct_toDate, file0, usecols, team, attention, emailbodytext, files, subject,  html_table1, html_table2, result1, result2)


    # Construct the file pattern to match
    file_pattern = '*SO*.xlsx'
    
    # Get a list of file paths that match the pattern
    file_paths = glob.glob(os.path.join(working_folder, file_pattern))
    
    # Iterate over the file paths and delete each file
    for file_path in file_paths:

        # File is automatically closed after exiting the with block
        # Remove the file
        os.remove(file_path)        
        
        

def email_proc(sender_email, receiv_email, cc_email, passwd_email, unstruct_toDate, 
               file0, usecols, team, attention, emailbodytext, files, subject, html_table1, html_table2, result1, result2):

    import pandas as pd
    from datetime import datetime, timedelta
    from email.message import EmailMessage
    import smtplib
    import os.path
    import ssl
    #from passwd import pswd
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from openpyxl import load_workbook
    import openpyxl
    
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiv_email
    msg['CC'] = cc_email
    msg['Subject'] = subject
    
    # Read the CSV file into a pandas DataFrame
    #df = pd.read_csv(file0)
    
    # Convert the pandas DataFrame to a CSV file
    #df.to_csv(file0, index=False)
    
    # Read CSV file and extract desired rows and columns
    #df = pd.read_csv(file0, usecols=usecols)
    #nrows = df.shape[0]
    
    # Read CSV file again using nrows
    #df = pd.read_csv(file0, usecols=usecols, nrows=nrows)

    # Load PivotTable from Excel file
    #workbook = openpyxl.load_workbook(file0)
    #worksheet = workbook['Pivot Table']
    #pivot_table = worksheet['A1'].pivotTableDefinition.cacheDefinition
    
    # Convert PivotTable to pandas DataFrame
    #df = pd.read_excel(file0, sheet_name='Pivot Table', index_col=0, skiprows=pivot_table.row_cache_start, nrows=pivot_table.row_count)
    
    # Generate HTML table from DataFrame
    #html_table = df.to_html()
    
    # Convert DataFrame to HTML table
    #html_table = df.to_html(index=False)
    
    # Convert DataFrame to styled HTML table
    #html_table = (df.style
    #    .set_properties(**{'border-collapse': 'collapse', 'border': '1px solid black'})
    #    .set_table_styles([{'selector': 'th', 'props': [('background-color', 'lightblue'),
    #                                                     ('font-weight', 'bold'),
    #                                                     ('border', '1px solid black')]},
    #                       {'selector': 'tr:last-child', 'props': [('background-color', 'lightblue'),
    #                                                                ('font-weight', 'bold'),
    #                                                                ('border', '1px solid black')]}])
    #    .set_table_attributes('border="1" class="dataframe table table-striped table-hover table-sm" style="text-align: center; font-family: Arial"')
    #    .set_table_attributes([{'selector': 'thead tr',
    #                            'props': [('background-color', 'lightblue')]}])
    #    .set_table_attributes([{'selector': 'tbody tr:nth-child(1), tbody tr:nth-child(2)',
    #                            'props': [('background-color', 'lightblue')]}])
    #    .render()
    #)
    
    # Add HTML message to email body
    html_content = f"""
    <html>
        <body>
            <p> </p>
            <p> Kind attention required from: OPS Team</p>
            <p> </p>
            <p> KPI for Order Download KPI</p>
            <p> Order Picked/Downloaded Before 11am and 2pm :-  <b>{result1} % Orders Downloaded after 11am and 2pm</b> <p>  </p>
            <p> - 100% of Eligible lines picked/downloaded before 11:00am EST (exclude holds, etc.)</p>
            <p> - 11am net-change runs, new lines downloaded by 2:00pm EST.</p>     
            <p> - Errors should be Less than 1% error                                :-  <b>{result2} % Errors</b> <p>
            <p> </p>
            <p>{emailbodytext}</p>
            <p> </p>
            <p>Parts Automation-Python</p>
            <p> </p>
        </body>
    </html>
    """
    # Add HTML table to email body
    html_body = MIMEText(html_content + html_table2 + html_table1, 'html')
    msg.attach(html_body)
    
    # Attach multiple files to email
    files = files
    
    # List of paths to files
    for file in files:
        with open(file, 'rb') as f:
            file_content = f.read()
            # Create attachment and add to message
            attachment = MIMEApplication(file_content, Name=file)
            attachment['Content-Disposition'] = f'attachment; filename="{file}"'
            msg.attach(attachment)
   
    # Send email
    #with smtplib.SMTP('smtp.office365.com', 587) as smtp:
    #    context = ssl.create_default_context()
    #    smtp.ehlo()
    #    smtp.starttls(context=context)
    #    smtp.login(sender_email, passwd_email)
    #    smtp.sendmail(sender_email, receiv_email.split(','), msg.as_string())
    with smtplib.SMTP('smtp.office365.com', 587) as smtp:
        context = ssl.create_default_context()
        smtp.ehlo()
        smtp.starttls(context=context)
        smtp.login(sender_email, passwd_email)
        receiv_email_list = receiv_email.split(',')
        cc_email_list = cc_email.split(',')
        receiv_email_list.extend(cc_email_list)
        smtp.sendmail(sender_email, receiv_email_list, msg.as_string())        

        print("job has been succesfully executed ", datetime.now())

   
def main():
    import sys
    import datetime
    sub_proc1()
    
if __name__ == "__main__":
    main()