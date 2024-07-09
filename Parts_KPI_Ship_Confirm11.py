def sub_proc1():
    import cx_Oracle as ora
    import pymysql
    import pandas as pd
    import os
    import glob
    import sys
    from datetime import datetime, timedelta
    import csv
    from email.message import EmailMessage
    import smtplib
    import os.path
    import ssl
    import numpy as np
    import openpyxl  
    import xlsxwriter
    #from passwd import pswd
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from cryptography.fernet import Fernet
    import configparser
    from collections import OrderedDict
    os.chdir("C:\Parts POSB Automation")
    # Load the encryption key
    with open('encryption_key.key', 'rb') as file:
        key = file.read()
    
    # Create a Fernet cipher using the key
    cipher_suite = Fernet(key)
    
    # Read the encrypted data from the encrypted_config.ini file
    with open('encrypted_config.ini', 'rb') as file:
        ciphertext = file.read()
    
    # Decrypt the ciphertext using the cipher
    plaintext = cipher_suite.decrypt(ciphertext)
    
    # Convert the plaintext to a string
    plaintext_str = plaintext.decode('utf-8')
    
    # Parse the decrypted config data into a dictionary
    config_data = configparser.ConfigParser(dict_type=OrderedDict)
    config_data.read_string(plaintext_str)
    
    credentials = config_data['Credentials']
    connstr_plsql_v = credentials['connstr_plsql']
    connstr_mysql_v = credentials['connstr_mysql']
    host_mysql_v = credentials['host_mysql']
    user_mysql_v = credentials['user_mysql']
    pswd_mysql_v = credentials['pswd_mysql']
    dbas_mysql_v = credentials['dbas_mysql']
    passwd_email = credentials['passwd_email']
    sender_email = credentials['sender_email']
    receiv_email = credentials['receiv_email']
    cc_email = credentials['cc_email']
    attention = credentials['attention0']
    working_folder = credentials['working_folder']
    os.chdir(working_folder)

    print(working_folder)
    
    conn=pymysql.connect(host=host_mysql_v, user=user_mysql_v,  passwd=pswd_mysql_v, database=dbas_mysql_v)
    
    ora.init_oracle_client(
        lib_dir=r"C:\cx_Oracle\instantclient-basic-windows.x64-21.8.0.0.0dbru\instantclient-basic-windows.x64-21.8.0.0.0dbru\instantclient_21_8")
    #connstr = 'RO999999999/xxxxxxxx@geaebsdbadg1.appl.ge.com:1521/ercebs1p'
    connection = ora.connect(connstr_plsql_v)
    # cur=connection.cursor()
    now = datetime.now()
    # Check if today is Monday
    if now.weekday() == 0:
        # If it is Monday, set unstruct_fromDate to two days ago
        unstruct_fromDate = now + timedelta(days=-3)
        unstruct_fromDate3 = now + timedelta(days=-10)                              
        fromDate3 = unstruct_fromDate3.strftime("%d-%b-%Y 00:00:01").upper()
        print("date and time =", fromDate3)
        unstruct_toDate3 = now + timedelta(days=-4)                              
        toDate3 = unstruct_toDate3.strftime("%d-%b-%Y 23:59:59").upper()
    else:
        # If it is not Monday, set unstruct_fromDate to one day ago
        unstruct_fromDate = now + timedelta(days=-1)
        unstruct_fromDate3 = now + timedelta(days=-8)                              
        fromDate3 = unstruct_fromDate3.strftime("%d-%b-%Y 00:00:01").upper()
        print("date and time =", fromDate3)
        unstruct_toDate3 = now + timedelta(days=-2)                              
        toDate3 = unstruct_toDate3.strftime("%d-%b-%Y 23:59:59").upper()
        
    unstruct_toDate = now                              
    toDate4 = unstruct_toDate.strftime("%d-%b-%Y 23:59:59").upper()
    print("future date =", toDate4)  
    
    unstruct_toDate = now + timedelta(days=-1)           
    fromDate = unstruct_fromDate.strftime("%d-%b-%Y 00:00:01").upper()
    fromDate2 = unstruct_fromDate.strftime("%Y-%m-%d 00:00:01").upper()
    print("date and time =", fromDate)
    print("date and time =", fromDate2)
    toDate = unstruct_toDate.strftime("%d-%b-%Y 23:59:59").upper()
    toDate2 = unstruct_toDate.strftime("%Y-%m-%d 23:59:59").upper()
    print("future date =", toDate)
    print("future date =", toDate2)
    toDate1 = unstruct_toDate.strftime("%d-%b-%Y")    
    print("future date =", toDate3)
    fromDate4 = unstruct_fromDate.strftime("%Y-%m-%d").upper()
    
     
    file1="Ship_Confirm_Stage_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file2="Ship_Confirm_BaseT_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file3="Ship_Confirm_Unmatched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file4="Ship_Confirm_Matched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file5="Ship_Confirm_Merged_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file6="Ship_Confirm_Errors_Overdue_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file7="Ship_Confirm_WMS_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file8="Ship_Confirm_Unmatched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file9="Ship_Confirm_Matched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file10="Ship_Confirm_Merged_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"


    print(file1)
    print(file2)
    print(file3)
    print(file4)
    print(file5)   
    print(file6)
    print(file7) 
    print(file8) 
    print(file9) 
    print(file10)  
    
    query_result1=("""
        Select 
         concat(sc.organization_code,sc.Item_number) key1
        ,concat(concat(sc.organization_code,sc.Item_number),to_char(sc.order_number)) key2
        ,concat(concat(concat(sc.organization_code,sc.Item_number),to_char(sc.order_number)),abs(sc.shipped_quantity)) KEY3
        ,sc.organization_code
        ,sc.Item_number
        ,sc.order_number
        ,sc.attribute3
        ,sc.attribute13 WMS_ORDER
        ,sc.attribute14
        ,sc.attribute18 ORDER_TYPE
        ,sc.attribute19
        ,sc.Shipped_quantity
        ,sc.shipment_id
        ,sc.shipped_date
        ,sc.creation_date
        ,trunc(sc.creation_date) STG_DATE
        ,sc.last_update_Date
        ,sc.delivery_number
        ,sc.creation_date-shipped_Date as time_lag_to_interface
        ,sc.last_update_date-creation_Date as time_lag_to_get_process_in_EBS
        ,null as RUN_TIME
        ,null as IN_EBS
        ,null as NOT_IN_EBS
        ,null as EXCEEDS_1_HR
        ,1 as TOTAL_COUNT
        ,sc.process_flag
        from apps.GEWSH_APL_SHIP_CONFIRM_TBL sc
        where 1=1
        and sc.process_flag != 'N'
        and sc.creation_date >= TO_DATE('"""+fromDate+"""', 'DD-MON-YYYY hh24:mi:ss')
        -- and sc.creation_date BETWEEN TO_DATE('"""+fromDate+"""', 'DD-MON-YYYY hh24:mi:ss')
        -- AND TO_DATE('"""+toDate+"""', 'DD-MON-YYYY hh24:mi:ss')
        and sc.line_status = 'NO'
        order by sc.shipped_date
    """)
    
    query_result2=("""
        SELECT
         concat(mp.organization_code,msi.segment1) key1
        ,concat(concat(mp.organization_code,msi.segment1),to_char(ooh1.order_number)) key2
        ,concat(concat(concat(mp.organization_code,msi.segment1),to_char(ooh1.order_number)),abs(mmt.transaction_quantity)) KEY3
        ,mp.organization_code
        ,msi.item_type item_type
        ,msi.segment1 item_number
        ,mmt.SHIPMENT_NUMBER
        ,msi.inventory_item_status_code
        ,mmt.source_line_id
        ,mmt.transaction_id
        ,mmt.transaction_type_id
        ,mtt.transaction_type_name
        --,MTA.reference_account
        --,GCC.concatenated_segments
        ,mmt.transaction_date 
        ,mmt.creation_date 
        ,trunc(mmt.creation_date) STG_DATE
        ,mmt.last_update_Date
        --,mmt.subinventory
        ,mmt.transaction_quantity
        --,mta.primary_quantity cnt
        --,mta.base_transaction_value
        ,mmt.last_updated_by
        ,mmt.created_by
        ,fu.user_name
        ,fu.description
        ,ooh.order_number RMA
        ,rt.subinventory
        ,rt.po_unit_price
        ,rt.quantity qty_received
        ,ooh1.order_number sales_order
        ,null as RUN_TIME
        ,null as IN_EBS
        ,null as NOT_IN_EBS
        ,null as EXCEEDS_1_HR
        ,1 as TOTAL_COUNT
        FROM
         apps.mtl_parameters mp
        ,apps.mtl_system_items_b msi
        ,apps.mtl_material_transactions mmt
        ,apps.rcv_transactions rt
        ,apps.oe_order_headers_all ooh
        ,apps.fnd_user fu
        ,apps.mtl_transaction_types mtt
        ,apps.wsh_delivery_details wdd
        ,apps.oe_order_lines_all ool
        ,apps.oe_order_headers_all ooh1
        --apps.GL_CODE_COMBINATIONS_KFV GCC
        --apps.MTL_TRANSACTION_ACCOUNTS MTA
        WHERE
        1=1
        AND mmt.transaction_type_id in (33,62) --in (31, 41) --in (10,13,12,15,16,18,19,22,27,1005,99,59,76,69,120,61,40,41,42,4) --= -- --in (31, 41) --in (31, 41,44,35) -= 44 --= 33 --
        AND mp.organization_code in ('NAP','WAL','JEF', 'RVR')
        AND mmt.inventory_item_id=msi.inventory_item_id
        AND mmt.organization_id=msi.organization_id
        AND mp.organization_id =msi.organization_id
        AND mmt.rcv_transaction_id=rt.transaction_id(+)
        AND rt.oe_order_header_id=ooh.header_id(+)
        --and mmt.transaction_id = mta.transaction_id
        --and mta.reference_account = gcc.code_combination_id
        --AND mmt.organization_id=mta.organization_id
        AND mmt.TRANSACTION_ACTION_ID =mmt.TRANSACTION_ACTION_ID+0
        AND mmt.TRANSACTION_SOURCE_TYPE_ID=mmt.TRANSACTION_SOURCE_TYPE_ID+0
        --AND mmt.transaction_date BETWEEN TO_DATE('17-JAN-2023 00:00:00', 'DD-MON-YYYY hh24:mi:ss')
        --AND TO_DATE('17-JAN-2023 23:59:59','DD-MON-YYYY hh24:mi:ss')
        AND mmt.last_update_date BETWEEN TO_DATE('"""+fromDate+"""', 'DD-MON-YYYY hh24:mi:ss') AND TO_DATE('"""+toDate4+"""', 'DD-MON-YYYY hh24:mi:ss')
        AND mmt.created_by=fu.user_id
        AND wdd.delivery_Detail_id(+) = mmt.picking_line_id
        and ool.line_id(+) = wdd.source_line_id
        and ooh1.header_id(+)=ool.header_id
        --AND wdd.source_code = 'OE'
        --and wdd.oe_interfaced_flag = 'Y'
        --and wdd.released_status = 'Y'
        --and mmt.last_updated_by in (32198 7 153740
        --and mmt.created_by in (32198,15370)
        and mtt.transaction_type_id = mmt.transaction_type_id
        --AND MMT.SOURCE_LINE_ID IN (271099929,85119,85121,85108)
        --and mmt.created_by NOT in (32198,75078, 153740,73977)
        --and msi.segment1 ='234D1704P001'
        --and transaction_type_name='Sales order issue'
        --and mmt.SHIPMENT_NUMBER='98811647'
        order by mmt.transaction_id
    """)
        
    query_result3=("""
        Select 
         concat(sc.organization_code,sc.Item_number) key1
        ,concat(concat(sc.organization_code,sc.Item_number),to_char(sc.order_number)) key2
        ,concat(concat(concat(sc.organization_code,sc.Item_number),to_char(sc.order_number)),abs(sc.shipped_quantity)) KEY3
        ,sc.organization_code WAREHOUSE
        ,sc.Item_number
        ,sc.order_number
        ,sc.attribute3
        ,sc.attribute13 WMS_ORDER
        ,sc.attribute14
        ,sc.attribute18 ORDER_TYPE
        ,sc.attribute19
        ,sc.Shipped_quantity
        ,sc.shipment_id
        ,sc.shipped_date
        ,sc.creation_date
        ,trunc(sc.creation_date) STG_DATE
        ,sc.last_update_Date
        ,sc.last_update_date-sc.creation_date REPROCESS_TIME
        ,sc.delivery_number
        ,sc.creation_date-shipped_Date as time_lag_to_interface
        ,sc.last_update_date-creation_Date as time_lag_to_get_process_in_EBS
        ,null as RUN_TIME
        ,null as IN_EBS
        ,null as NOT_IN_EBS
        ,null as EXCEEDS_1_HR
        ,1 as TOTAL_COUNT
        ,sc.process_flag
        ,sc.line_status
        ,1 as ERRORS_PENDING_TO_REPROCESS
        from apps.GEWSH_APL_SHIP_CONFIRM_TBL sc
        where 1=1
        and sc.process_flag = 'E'
        and sc.creation_date >= TO_DATE('"""+fromDate3+"""', 'DD-MON-YYYY hh24:mi:ss')
        and sc.line_status = 'NO'
        order by sc.shipped_date
    """)

    query_result4=("""
    	select 
         concat(warehouse,item) KEY1
        ,concat(concat(warehouse,item),SUBSTRING_INDEX(ERP_ORDER,'-',1)) KEY2
        ,concat(concat(concat(warehouse,item),SUBSTRING_INDEX(ERP_ORDER,'-',1)),ROUND(abs(TOTAL_QTY),0)) KEY3
        ,WAREHOUSE
        ,SHIPMENT_ID shipment_num
        ,ERP_ORDER
        ,item
        ,TOTAL_QTY  QUANTITY
        ,DATE_TIME_STAMP ACTIVITY_DATE_TIME
        ,DATE_FORMAT(DATE_TIME_STAMP, '%Y-%m-%d') WMS_ACTIVITY_DATE_TIME
        ,NULL IN_STAGE
        ,NULL IN_EBS
        ,NULL NOT_IN_STAGE
        ,NULL NOT_IN_EBS
        ,NULL RUN_TIME
        ,NULL EXCEEDS_1_HR
        ,NULL STG_CREATION_DATE
        ,NULL MTL_CREATION_DATE
        ,1 TOTAL_COUNT
        ,null REPROCESS_COUNT
        from upload_order_detail
        where INTERFACE_CONDITION='Processed'
        and DATE_TIME_STAMP >= '"""+fromDate2+"""'
        and DATE_TIME_STAMP <= '"""+toDate2+"""'       
        and SHIPPED_QTY <>0
    """)
    
    query_result5=("""
        select 
         concat(warehouse,item) KEY1
        ,concat(concat(warehouse,item),left(ERP_ORDER_number,10)) KEY2
        ,concat(concat(concat(warehouse,item),left(ERP_ORDER_number,10)),abs(quantity)) KEY3
        ,warehouse WAREHOUSE
        ,internal_shipment_num shipment_num
        ,erp_order_number ERP_ORDER
        ,item
        ,quantity QUANTITY
        ,date_and_time ACTIVITY_DATE_TIME
        ,DATE_FORMAT(date_and_time, '%Y-%m-%d') WMS_ACTIVITY_DATE_TIME     
        ,NULL IN_STAGE
        ,NULL IN_EBS
        ,NULL NOT_IN_STAGE
        ,NULL NOT_IN_EBS
        ,NULL RUN_TIME
        ,NULL EXCEEDS_1_HR
        ,NULL STG_CREATION_DATE
        ,NULL MTL_CREATION_DATE
        ,1 TOTAL_COUNT
        ,null REPROCESS_COUNT
        from irms_data.ship_confirmations_cons
    	where 
    	DATE_FORMAT(date_and_time, '%Y-%m-%d hh24:mi:ss') >= '"""+fromDate2+"""' and DATE_FORMAT(date_and_time, '%Y-%m-%d hh24:mi:ss') <= '"""+toDate2+"""'
    	and quantity <>0
    	and item<>'0'
    """)

    df4=pd.read_sql_query(query_result4,conn)
    header=list(df4.columns)
    print(df4.columns)
    print(file7)
    df5=pd.read_sql_query(query_result5,conn)
    header=list(df5.columns)
    print(df5.columns)
    combined_df = pd.concat([df4, df5], ignore_index=True)
    writer = pd.ExcelWriter(file7, engine='openpyxl')
    combined_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()

    df=pd.read_sql_query(query_result3,connection)
    header=list(df.columns)
    print(file6)
    writer = pd.ExcelWriter(file6, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    df_pending = df
    
    df=pd.read_sql_query(query_result1,connection)
    header=list(df.columns)
    print(file1)
    writer = pd.ExcelWriter(file1, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    df=pd.read_sql_query(query_result2,connection)
    header=list(df.columns)
    print(file2)
    writer = pd.ExcelWriter(file2, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    key1 = 'KEY2'
    key2 = 'KEY2'
    key3 = 'KEY2'   
    
    # Use pandas to read the Excel file into a DataFrame
    df1 = pd.read_excel(file1, sheet_name='Sheet1')
    df2 = pd.read_excel(file2, sheet_name='Sheet1')
    df3 = pd.read_excel(file7, sheet_name='Sheet1')
    
    # Find unmatched & matched records from scale to stage
    scale_unmatched_records = []
    scale_matched_records = []
    
    for index, row in df3.iterrows():
        if row[key3] not in df1[key1].values and row[key3] not in df2[key2].values:
            scale_unmatched_record = row.to_dict()
            scale_unmatched_record['RUN_TIME'] = pd.Timedelta(0)
            scale_unmatched_record['IN_STAGE'] = 'NOT_IN_STAGE'
            scale_unmatched_record['NOT_IN_STAGE'] = 1
            scale_unmatched_record['IN_EBS'] = 'NOT_IN_EBS'
            scale_unmatched_record['NOT_IN_EBS'] = 1
            scale_unmatched_record['EXCEEDS_1_HR'] = 0  
            scale_unmatched_records.append(scale_unmatched_record)
            
        elif row[key3] in df1[key1].values and row[key3] in df2[key2].values:
            scale_matching_row1 = df1[df1[key1] == row[key3]]
            scale_matching_row2 = df2[df2[key2] == row[key3]]
            scale_matched_record = row.to_dict()
            scale_matched_record['STG_CREATION_DATE'] = scale_matching_row1['CREATION_DATE'].values[0]
            scale_matched_record['MTL_CREATION_DATE'] = scale_matching_row2['CREATION_DATE'].values[0]
            scale_matched_record['RUN_TIME'] = scale_matching_row2['CREATION_DATE'].values[0].astype('datetime64[ns]') - row['ACTIVITY_DATE_TIME']
            scale_matched_record['IN_STAGE'] = 'IN_STAGE'
            scale_matched_record['NOT_IN_STAGE'] = 0
            scale_matched_record['IN_EBS'] = 'IN_EBS'
            scale_matched_record['NOT_IN_EBS'] = 0
            scale_matched_record['EXCEEDS_1_HR'] = int(scale_matched_record['RUN_TIME'] > pd.Timedelta(hours=4))
            scale_matched_records.append(scale_matched_record)
            
        elif row[key3] in df1[key1].values and row[key3] not in df2[key2].values:
            scale_matching_row1 = df1[df1[key1] == row[key3]]
            scale_unmatched_record = row.to_dict()
            scale_unmatched_record['STG_CREATION_DATE'] = scale_matching_row1['CREATION_DATE'].values[0]
            scale_unmatched_record['RUN_TIME'] = pd.Timedelta(0)
            scale_unmatched_record['IN_STAGE'] = 'IN_STAGE'
            scale_unmatched_record['NOT_IN_STAGE'] = 0
            scale_unmatched_record['IN_EBS'] = 'NOT_IN_EBS'
            scale_unmatched_record['NOT_IN_EBS'] = 1
            scale_unmatched_record['EXCEEDS_1_HR'] = 0
            scale_unmatched_records.append(scale_unmatched_record)
            
        elif row[key3] not in df1[key1].values and row[key3] in df2[key2].values:
            scale_matching_row2 = df2[df2[key2] == row[key3]]
            scale_unmatched_record = row.to_dict()
            scale_unmatched_record['MTL_CREATION_DATE'] = scale_matching_row2['CREATION_DATE'].values[0]
            scale_unmatched_record['RUN_TIME'] = scale_matching_row2['CREATION_DATE'].values[0].astype('datetime64[ns]') - row['ACTIVITY_DATE_TIME']
            scale_unmatched_record['IN_STAGE'] = 'NOT_IN_STAGE'
            scale_unmatched_record['NOT_IN_STAGE'] = 1
            scale_unmatched_record['IN_EBS'] = 'IN_EBS'
            scale_unmatched_record['NOT_IN_EBS'] = 0
            scale_unmatched_record['EXCEEDS_1_HR'] = int(scale_unmatched_record['RUN_TIME'] > pd.Timedelta(hours=4))
            scale_unmatched_records.append(scale_unmatched_record)
    
    # Convert the lists of unmatched and matched records to DataFrames
    scale_unmatched_records_df = pd.DataFrame(scale_unmatched_records)
    scale_matched_records_df = pd.DataFrame(scale_matched_records)
   
    header=list(scale_unmatched_records_df.columns)
    print(file8)
    writer = pd.ExcelWriter(file8, engine='openpyxl')
    scale_unmatched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    header=list(scale_matched_records_df.columns)
    print(file9)
    writer = pd.ExcelWriter(file9, engine='openpyxl')
    scale_matched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    # Combine the scale_unmatched and matched records
    result_scale_df = pd.concat([scale_unmatched_records_df, scale_matched_records_df], axis=0, ignore_index=True)
    #result_scale_df = result_scale_df.sort_values(by=['WMS_ACTIVITY_DATE_TIME'])

    # Create a new Excel file with the combined results
    with pd.ExcelWriter(file10, engine='xlsxwriter') as writer:
        # Write the result DataFrame to a new sheet
        result_scale_df.to_excel(writer, sheet_name='Sheet1', index=False)
    
    # Format the 'Unmatched' and 'Matched' columns
    worksheet = writer.sheets['Sheet1']
    scale_unmatched_cond_fmt = {
        'type': 'cell',
        'criteria': 'equal to',
        'value': '"Unmatched"',
        'format': writer.book.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    }
    scale_matched_cond_fmt = {
        'type': 'cell',
        'criteria': 'equal to',
        'value': '"Matched"',
        'format': writer.book.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
    }
    worksheet.conditional_format('G2:G{}'.format(len(result_scale_df)+1), scale_unmatched_cond_fmt)
    worksheet.conditional_format('G2:G{}'.format(len(result_scale_df)+1), scale_matched_cond_fmt)

            
    # Convert the lists of unmatched and matched records to DataFrames
    #unmatched_records_df = pd.DataFrame(unmatched_records)
    #matched_records_df = pd.DataFrame(matched_records)
   
    #header=list(unmatched_records_df.columns)
    #print(file3)
    #writer = pd.ExcelWriter(file3, engine='openpyxl')
    #unmatched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    #writer.close()

    #header=list(matched_records_df.columns)
    #print(file4)
    #writer = pd.ExcelWriter(file4, engine='openpyxl')
    #matched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    #writer.close()
        
    # Combine the unmatched and matched records
    #result_df = pd.concat([unmatched_records_df, matched_records_df], axis=0, ignore_index=True)
    #result_df = result_df.sort_values(by=['SHIP_CONFIRM_IN_STAGE', 'IN_EBS', 'PROCESS_FLAG' ,'ORDER_TYPE'])
    
    # Create a new Excel file with the combined results
    #with pd.ExcelWriter(file5, engine='xlsxwriter') as writer:
        # Write the result DataFrame to a new sheet
    #    result_df.to_excel(writer, sheet_name='Combined', index=False)
    
        # Format the 'Unmatched' and 'Matched' columns
    #    worksheet = writer.sheets['Combined']
    #    unmatched_cond_fmt = {
    #        'type': 'cell',
    #        'criteria': 'equal to',
    #        'value': '"Unmatched"',
    #        'format': writer.book.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    #    }
    #    matched_cond_fmt = {
    #        'type': 'cell',
    #        'criteria': 'equal to',
    #        'value': '"Matched"',
    #        'format': writer.book.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
    #    }
    #    worksheet.conditional_format('G2:G{}'.format(len(result_df)+1), unmatched_cond_fmt)
    #    worksheet.conditional_format('G2:G{}'.format(len(result_df)+1), matched_cond_fmt)

    # Reorder the columns in result_scale_df
    if not result_scale_df.empty:
        result_scale_df_ordered = result_scale_df[['WMS_ACTIVITY_DATE_TIME', 'WAREHOUSE', 'IN_STAGE', 'IN_EBS', 'TOTAL_COUNT', 'NOT_IN_STAGE', 'NOT_IN_EBS', 'EXCEEDS_1_HR', 'RUN_TIME', 'QUANTITY']]
    
        # Create an OrderedDict with the desired column order
        aggfunc_dict = OrderedDict([
            ('TOTAL_COUNT', 'sum'),
            ('NOT_IN_STAGE', 'sum'),
            ('NOT_IN_EBS', 'sum'),
            ('EXCEEDS_1_HR', 'sum'),
            ('RUN_TIME', [np.max, np.mean]),
            ('QUANTITY', 'sum')
        ])
        
       
        # Perform the pivot operation using the ordered columns
        pt1 = pd.pivot_table(result_scale_df_ordered.reset_index(),
                             index=['WMS_ACTIVITY_DATE_TIME', 'WAREHOUSE', 'IN_STAGE', 'IN_EBS'],
                             aggfunc=aggfunc_dict)

        pt1 = pt1.rename(columns={'EXCEEDS_1_HR': 'EXCEEDS_4_HRS'})    
            
        # Reorder the columns based on the desired column order
        #pt1 = pt1.reindex(columns=list(aggfunc_dict.keys()))
    
        # Rename columns in MultiIndex
        pt1.columns = list(map('_'.join, pt1.columns.values))
    
        # Convert PivotTable to HTML table
        html_table4 = pt1.to_html()
    else:
        html_table4 = ''

    # Define the order of the columns
    #column_order = ['SHIP_CONFIRM_IN_STAGE', 'IN_EBS', 'PROCESS_FLAG' ,'ORDER_TYPE']
       
    # Create a pivot table
    #pt = pd.pivot_table(result_df.reset_index(),
    #                index=['SHIP_CONFIRM_IN_STAGE', 'IN_EBS', 'PROCESS_FLAG' ,'ORDER_TYPE'],
    #                aggfunc={'TOTAL_COUNT':'sum',
    #                         'SHIPPED_QUANTITY':'sum',
    #                         'NOT_IN_EBS': 'sum',
    #                         'EXCEEDS_1_HR': 'sum',
    #                         'RUN_TIME': [np.mean,np.max]})

    #pt = pt.round(0)  # Round the values to 2 decimal places

    # Combine the defect percentage with the main pivot table
    #pt = pt.merge(defect_percent['DEFECT_PERCENT'], left_on='SHIP_CONFIRM_IN_STAGE', right_on='SHIP_CONFIRM_IN_STAGE').round(0)
    
    # Rename columns in MultiIndex
    #pt.columns = list(map('_'.join, pt.columns.values))

    # Define a function to highlight based on condition
    #def highlight_col(col):
    #    if col['NOT_IN_EBS'] > 60:
    #        return ['background-color: yellow']
    #    else:
    #        return ['background-color: white']
    
    # Define a function to highlight rows based on condition
    #def highlight_defects(row):
    #    if (row['SOA_STATUS'] == 'ERROR') or (row['SOA_STATUS'] == 'STUCK'):
    #        return ['background-color: red']
    #    else:
    #        return ['background-color: white']
        
    # Apply the style to the pivot table
    #styled_pt = pt.style.applymap(highlight_col, subset=pd.IndexSlice[:, :])
    # Apply the style to the pivot table
    #styled_pt = pt.style.applymap(highlight_defects, subset=pd.IndexSlice[:, :])
    
    # Convert PivotTable to HTML table
    #html_table2 = pt.to_html()

    # Calculate the defect percentage for each date
    defect_percent = result_scale_df.groupby('WMS_ACTIVITY_DATE_TIME')[['NOT_IN_STAGE','NOT_IN_EBS', 'EXCEEDS_1_HR', 'TOTAL_COUNT']].sum()
     
    # Convert the 'NOT_IN_EBS' and 'RECEIPT_COUNT' columns to numeric data types
    defect_percent['NOT_IN_STAGE'] = pd.to_numeric(defect_percent['NOT_IN_STAGE'], errors='coerce')
    defect_percent['NOT_IN_EBS'] = pd.to_numeric(defect_percent['NOT_IN_EBS'], errors='coerce')
    defect_percent['EXCEEDS_1_HR'] = pd.to_numeric(defect_percent['EXCEEDS_1_HR'], errors='coerce')
    defect_percent['TOTAL_COUNT'] = pd.to_numeric(defect_percent['TOTAL_COUNT'], errors='coerce')
    
    # Drop any rows with missing or non-numeric values
    defect_percent.dropna(inplace=True)
     
    # Calculate the defect percentage for each date
    defect_percent['PERCENTAGE_NOT_IN_STAGE'] = defect_percent['NOT_IN_STAGE'] / defect_percent['TOTAL_COUNT'] * 100
    defect_percent['Errors < 1 %'] = defect_percent['NOT_IN_EBS'] / defect_percent['TOTAL_COUNT'] * 100
    defect_percent['PERCENTAGE_EXCEEDS_1_HR'] = defect_percent['EXCEEDS_1_HR'] / defect_percent['TOTAL_COUNT'] * 100
    
    result1 = (defect_percent['EXCEEDS_1_HR'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]
    result2 = (defect_percent['NOT_IN_EBS'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]
    result3 = (defect_percent['NOT_IN_STAGE'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]

    result1 = float(result1)
    result1 = round(result1,2)
    result2 = float(result2)
    result2 = round(result2,2)
    result3 = float(result3)
    result3 = round(result3,2)

    defect_percent = defect_percent.rename(columns={'EXCEEDS_1_HR': 'EXCEEDS_PROCESS_TIME', 'PERCENTAGE_EXCEEDS_1_HR': 'Exceeds Process Time'})
    #defect_percent = defect_percent.reset_index().rename(columns={'WMS_ACTIVITY_DATE_TIME': 'WMS_RECEIPT_DATE'})
    print(defect_percent.columns)    
     
    # Convert the defect percent data frame to an HTML table string
    html_table1 = defect_percent.to_html(index=True)
    
    #    defect_pending = df_pending.groupby('RECEIPT_IN_STAGE','WAREHOUSE').agg(ERRORS_PENDING_TO_REPROCESS=('ERRORS_PENDING_TO_REPROCESS', 'sum','REPROCESS_TIME', 'sum')).reset_index()
    defect_pending = df_pending.groupby(['STG_DATE', 'WAREHOUSE','PROCESS_FLAG']).agg({
        'ERRORS_PENDING_TO_REPROCESS': 'sum',
        'REPROCESS_TIME': 'sum'
    }).reset_index().rename(columns={
        'ERRORS_PENDING_TO_REPROCESS': 'ERRORS_PENDING_TO_REPROCESS_SUM',
        'REPROCESS_TIME': 'REPROCESS_TIME_SUM'
    })
    
    if not df_pending.empty:
        defect_percent['Errors not resolved within 24 hours'] = df_pending.groupby('PROCESS_FLAG').agg(ERRORS_PENDING_TO_REPROCESS=('ERRORS_PENDING_TO_REPROCESS', 'sum')).reset_index().iloc[0]['ERRORS_PENDING_TO_REPROCESS']
    else:
        defect_percent['Errors not resolved within 24 hours'] = 0    

    #defect_pending['ERRORS_PENDING_TO_REPROCESS'] = pd.to_numeric(defect_pending['ERRORS_PENDING_TO_REPROCESS'], errors='coerce') 
    defect_pending.dropna(inplace=True)
    if df_pending.empty:
        result4 = 0
    else:
        result4 = df_pending.groupby('PROCESS_FLAG').agg(ERRORS_PENDING_TO_REPROCESS=('ERRORS_PENDING_TO_REPROCESS', 'sum')).reset_index().iloc[0]['ERRORS_PENDING_TO_REPROCESS']
    
    # Convert the DataFrame to an HTML table
    html_table3 = defect_pending.to_html(index=False)
       
    # Open the Excel file and create an ExcelWriter object
    #writer = pd.ExcelWriter(file5, engine='openpyxl')
    #writer.book = openpyxl.load_workbook(file5)
    
    # Write the pivot table to a new sheet in the existing workbook
    #pt.to_excel(writer, sheet_name='Pivot Table')
    
    # Save the changes and close the workbook
    #writer.close()
    #writer.close()
  
    # Load the existing Excel file into a pandas DataFrame
    df_KPI = pd.read_excel('Parts_KPIs.xlsx')
    
    # Add custom text column to the pivot table DataFrame
    text_value = 'Ship Confirmations'
    defect_percent.insert(0, 'Parts_KPI', text_value)
    
    # Convert pivot table rows into a DataFrame
    df_append = pd.DataFrame(defect_percent.to_records())
    
    # Iterate over rows in df_append
    matching = 0
    for index, row in df_KPI.iterrows():
        #fromDate4 = row['WMS_ACTIVITY_DATE_TIME']
        existing_record1 = (df_append['Parts_KPI'] == text_value) & (df_append['WMS_ACTIVITY_DATE_TIME'] == fromDate4)
        #existing_record2 = (row['Parts_KPI'] == text_value) & (datetime.strptime(row['WMS_ACTIVITY_DATE_TIME'], "%Y-%m-%d").strftime("%Y-%m-%d") == fromDate4)
        existing_record2 = (row['Parts_KPI'] == text_value) & (row['WMS_ACTIVITY_DATE_TIME'] == fromDate4)
        
        # Compare with index record in df_KPI
        if len(df_append[existing_record1]) > 0:
            df_KPI_record = df_KPI.loc[index]  # Get the record at the current index in df_KPI
            append_record = df_append.loc[existing_record1].iloc[0]  # Get the matching record in df_append
            
            # Compare the record values
            if df_KPI_record.equals(append_record):
                matching = 1
                # Perform any desired actions or logic for matching records
    
        # Compare with index record in df_KPI
        if existing_record2:
            #if row['Parts_KPI'] == text_value and row['WMS_ACTIVITY_DATE_TIME'].strftime("%Y-%m-%d").upper() == fromDate4:
            if row['Parts_KPI'] == text_value and row['WMS_ACTIVITY_DATE_TIME'] == fromDate4:
                # Replace existing record with df_append record
                df_KPI.loc[index] = df_append.loc[existing_record1].iloc[0]
                matching = 1
    
        # Perform other operations or comparisons as needed
    
    if matching < 1:
        # Append new record
        df_KPI = pd.concat([df_KPI, df_append], ignore_index=False)
        
    # Convert 'Exceeds Process Time' column to numeric
    #df_KPI['Exceeds Process Time'] = pd.to_numeric(df_KPI['Exceeds Process Time'], errors='coerce')
    
    # Apply formatting to 'Exceeds Process Time' column
    #df_KPI['Exceeds Process Time'] = df_KPI['Exceeds Process Time'].apply(lambda x: "{:.1f}%".format(x) if pd.notnull(x) else "")        
    
    # Save the updated DataFrame to a new Excel file
    df_KPI.to_excel('Parts_KPIs.xlsx', index=False)
    
    file0=file5
    usecols= []
    team = 'OPS Team'
    subject = "Parts_KPI_Ship_Confirmation_{}"+toDate1
    emailbodytext = 'Please find attached Details for the Parts KPI Ship_Confirmation,  Request you to Report & Clear the  Defects Which are highlighted.'
    #files = [file1,file2,file6,file7,file8,file9,file10]
    files = [file8]
    
    email_proc(sender_email, receiv_email, cc_email, passwd_email, unstruct_toDate, file0, usecols, team, attention, emailbodytext, files, subject, html_table1, html_table3, html_table4, result1, result2, result3, result4)

    # Construct the file pattern to match
    file_pattern = 'Ship_Confirm*.xlsx'
    
    # Get a list of file paths that match the pattern
    file_paths = glob.glob(os.path.join(working_folder, file_pattern))
    
    # Iterate over the file paths and delete each file
    for file_path in file_paths:
        os.remove(file_path)


def email_proc(sender_email, receiv_email, cc_email, passwd_email, unstruct_toDate, file0, usecols, team, attention, emailbodytext, files, subject, html_table1, html_table3, html_table4, result1, result2, result3, result4):

    import pandas as pd
    from datetime import datetime, timedelta
    from email.message import EmailMessage
    import smtplib
    import os.path
    import ssl
    #from passwd import pswd
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from openpyxl import load_workbook
    import openpyxl
    
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiv_email
    msg['CC'] = cc_email
    msg['Subject'] = subject
    
    # Read the CSV file into a pandas DataFrame
    #df = pd.read_csv(file0)
    
    # Convert the pandas DataFrame to a CSV file
    #df.to_csv(file0, index=False)
    
    # Read CSV file and extract desired rows and columns
    #df = pd.read_csv(file0, usecols=usecols)
    #nrows = df.shape[0]
    
    # Read CSV file again using nrows
    #df = pd.read_csv(file0, usecols=usecols, nrows=nrows)

    # Load PivotTable from Excel file
    #workbook = openpyxl.load_workbook(file0)
    #worksheet = workbook['Pivot Table']
    #pivot_table = worksheet['A1'].pivotTableDefinition.cacheDefinition
    
    # Convert PivotTable to pandas DataFrame
    #df = pd.read_excel(file0, sheet_name='Pivot Table', index_col=0, skiprows=pivot_table.row_cache_start, nrows=pivot_table.row_count)
    
    # Generate HTML table from DataFrame
    #html_table = df.to_html()
    
    # Convert DataFrame to HTML table
    #html_table = df.to_html(index=False)
    
    # Convert DataFrame to styled HTML table
    #html_table = (df.style
    #    .set_properties(**{'border-collapse': 'collapse', 'border': '1px solid black'})
    #    .set_table_styles([{'selector': 'th', 'props': [('background-color', 'lightblue'),
    #                                                     ('font-weight', 'bold'),
    #                                                     ('border', '1px solid black')]},
    #                       {'selector': 'tr:last-child', 'props': [('background-color', 'lightblue'),
    #                                                                ('font-weight', 'bold'),
    #                                                                ('border', '1px solid black')]}])
    #    .set_table_attributes('border="1" class="dataframe table table-striped table-hover table-sm" style="text-align: center; font-family: Arial"')
    #    .set_table_attributes([{'selector': 'thead tr',
    #                            'props': [('background-color', 'lightblue')]}])
    #    .set_table_attributes([{'selector': 'tbody tr:nth-child(1), tbody tr:nth-child(2)',
    #                            'props': [('background-color', 'lightblue')]}])
    #    .render()
    #)
    
    # Add HTML message to email body
    # <p>  1. Ship Confirmations should get processed within 4 hours (Scale to EBS) :-  <b>{result1} % Ship Confirmations are Processed after 4 hours</b> <p>  Removed on 03/07/2024 on request from Bill St.Croix


    html_content = f"""
    <html>
        <body>
            <p> </p>
            <p> Kind attention required from: OPS Team</p>
            <p> </p>
            <p>KPI For Ship Confirmations<p>
            <p>  1. Ship Confirmations should get processed within 4 hours (Scale to EBS) :-  <b>{result1} % Ship Confirmations are Processed after 4 hours</b> <p>
             <p>  2. Errors should be Less than 1% error                                      :-  <b>{result2} % Errors</b> <p>
            <p>  3. Errors should resolve within 24 hours                                   :-  <b>{result4} Record/s of Ship Confirmation/s is/are Pending to reprocess by Ops Team</b>  <p>
            <p> </p>
            <p>{emailbodytext}</p>
            <p> </p>
            <p>Parts Automation-Python</p>
            <p> </p>
        </body>
    </html>
    """
    
    
    # Add HTML table to email body
    html_body = MIMEText(html_content + html_table4  + html_table1 + html_table3, 'html')
    msg.attach(html_body)
    
    # Attach multiple files to email
    files = files
    
    # List of paths to files
    for file in files:
        with open(file, 'rb') as f:
            file_content = f.read()
            # Create attachment and add to message
            attachment = MIMEApplication(file_content, Name=file)
            attachment['Content-Disposition'] = f'attachment; filename="{file}"'
            msg.attach(attachment)
    
    # Send email
    #with smtplib.SMTP('smtp.office365.com', 587) as smtp:
    #    context = ssl.create_default_context()
    #    smtp.ehlo()
    #    smtp.starttls(context=context)
    #    smtp.login(sender_email, passwd_email)
    #    smtp.sendmail(sender_email, receiv_email.split(','), msg.as_string())
    with smtplib.SMTP('smtp.office365.com', 587) as smtp:
        context = ssl.create_default_context()
        smtp.ehlo()
        smtp.starttls(context=context)
        smtp.login(sender_email, passwd_email)
        receiv_email_list = receiv_email.split(',')
        cc_email_list = cc_email.split(',')
        receiv_email_list.extend(cc_email_list)
        smtp.sendmail(sender_email, receiv_email_list, msg.as_string())        

        print("job has been succesfully executed ", datetime.now())

   
def main():
    import sys
    import datetime
    sub_proc1()
    
if __name__ == "__main__":
    main()