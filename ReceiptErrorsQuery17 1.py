def sub_proc1():
    import cx_Oracle as ora
    import pymysql
    import pandas as pd
    import os
    import glob
    import sys
    from datetime import datetime, timedelta
    import csv
    from email.message import EmailMessage
    import smtplib
    import os.path
    import ssl
    #from passwd import pswd
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    import configparser
    import sys ; sys.setrecursionlimit(sys.getrecursionlimit() * 5)
    from sqlalchemy import create_engine
    from collections import OrderedDict
    import numpy as np
    import openpyxl  
    import xlsxwriter
    from openpyxl import load_workbook 
    from openpyxl.workbook import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    from cryptography.fernet import Fernet
    import configparser
    from collections import OrderedDict
    os.chdir("C:\Parts POSB Automation")
    # Load the encryption key
    with open('encryption_key.key', 'rb') as file:
        key = file.read()
    
    # Create a Fernet cipher using the key
    cipher_suite = Fernet(key)
    
    # Read the encrypted data from the encrypted_config.ini file
    with open('encrypted_config.ini', 'rb') as file:
        ciphertext = file.read()
    
    # Decrypt the ciphertext using the cipher
    plaintext = cipher_suite.decrypt(ciphertext)
    
    # Convert the plaintext to a string
    plaintext_str = plaintext.decode('utf-8')
    
    # Parse the decrypted config data into a dictionary
    config_data = configparser.ConfigParser(dict_type=OrderedDict)
    config_data.read_string(plaintext_str)
    
    credentials = config_data['Credentials']
    connstr_plsql_v = credentials['connstr_plsql']
    connstr_mysql_v = credentials['connstr_mysql']
    host_mysql_v = credentials['host_mysql']
    user_mysql_v = credentials['user_mysql']
    pswd_mysql_v = credentials['pswd_mysql']
    dbas_mysql_v = credentials['dbas_mysql']
    passwd_email = credentials['passwd_email']
    sender_email = credentials['sender_email']
    receiv_email = credentials['receiv_email']
    cc_email = credentials['cc_email']
    attention = credentials['attention0']
    working_folder = credentials['working_folder']
    os.chdir(working_folder)

    print(working_folder)
    
    conn=pymysql.connect(host=host_mysql_v, user=user_mysql_v,  passwd=pswd_mysql_v, database=dbas_mysql_v)
    
    ora.init_oracle_client(lib_dir = r"C:\cx_Oracle\instantclient-basic-windows.x64-21.8.0.0.0dbru\instantclient-basic-windows.x64-21.8.0.0.0dbru\instantclient_21_8")
    #connstr = 'RO999999999/xxxxxxxx@geaebsdbadg1.appl.ge.com:1521/ercebs1p'
    connection = ora.connect(connstr_plsql_v)
    #engine = create_engine(connstr_plsql_v)
    
    #from sqlalchemy import create_engine
    #engine = create_engine(connection, connect_args={'connect_timeout': 300})

    # cur=connection.cursor()
    now = datetime.now()
    # Check if today is Monday
    if now.weekday() == 0:
        # If it is Monday, set unstruct_fromDate to two days ago
        unstruct_fromDate = now + timedelta(days=-2)
    else:
        # If it is not Monday, set unstruct_fromDate to one day ago
        unstruct_fromDate = now + timedelta(days=-40)
    fromDate = unstruct_fromDate.strftime("%d-%b-%Y 00:00:01").upper()
    fromDate2 = unstruct_fromDate.strftime("%Y-%m-%d 00:00:01").upper()
    print("date and time =", fromDate)
    print("date and time =", fromDate2)
    
    unstruct_toDate = now                             
    toDate4 = unstruct_toDate.strftime("%d-%b-%Y 23:59:59").upper()
    print("future date =", toDate4)    
    
    unstruct_toDate = now + timedelta(days=-1)                              
    toDate = unstruct_toDate.strftime("%d-%b-%Y 23:59:59").upper()
    toDate2 = unstruct_toDate.strftime("%Y-%m-%d 23:59:59").upper()
    print("future date =", toDate)
    print("future date =", toDate2)
    toDate1 = unstruct_toDate.strftime("%d-%b-%Y")    
    unstruct_fromDate3 = now + timedelta(days=-8)                              
    fromDate3 = unstruct_fromDate3.strftime("%d-%b-%Y 00:00:01").upper()
    print("date and time =", fromDate3)
    unstruct_toDate3 = now + timedelta(days=-1)                              
    toDate3 = unstruct_toDate3.strftime("%d-%b-%Y 23:59:59").upper()
    print("future date =", toDate3)
    fromDate4 = unstruct_fromDate.strftime("%Y-%m-%d").upper()

    
    file1="EBS_Receipts_Stage_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file2="EBS_Receipts_BaseT_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file3="EBS_Receipts_Error_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file4="EBS_Receipts_Match_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    print(file1)
    print(file2)
    print(file3)
    print(file4)
    file5="EBS_Receipts_Directly_Created_in_EBS_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"    
    file6="EBS_Receipts_Errors_Overdue_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file7="WMS_Receipts_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file8="WMS_Receipts_Errors_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file9="WMS_Receipts_Matched_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file10="WMS_Receipts_Merged_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    file11="EBS_Receipts_Stage_Errors_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())+".xlsx"
    print(file5)
    print(file6) 
    print(file7) 
    print(file8) 
    print(file9) 
    print(file10)
    print(file11)
    
    query_result1=("""
    select
    Concat(rh.rcv_organization_code,rl.item_number) key1,
    -- ,Concat(Concat(rh.rcv_organization_code,rl.item_number),decode(rh.transaction_type,'RMA',rl.attribute14,rh.shipment_num)) KEY2  Mangesh 02292024
    -- CASE WHEN rh.rcv_organization_code in ('JEF','RVR') THEN Concat('UID.',trim(rl.container_number))
    --     WHEN rh.rcv_organization_code in ('NAP','WAL') THEN Concat(Concat(rh.rcv_organization_code,rl.item_number),decode(rh.transaction_type,'RMA',rl.attribute14,rh.shipment_num))
    CASE WHEN rh.rcv_organization_code in ('JEF','RVR') THEN concat('UID.',trim(rl.container_number))
         WHEN rh.rcv_organization_code in ('NAP','WAL') THEN Concat(Concat(rh.rcv_organization_code,rl.item_number),decode(rh.transaction_type,'RMA',rl.attribute14,rh.shipment_num))
    END KEY2,
     Concat(Concat(Concat(rh.rcv_organization_code,rl.item_number),decode(rh.transaction_type,'RMA',rl.attribute14,rh.shipment_num)),rl.quantity) KEY3
    ,rh.transaction_type
    ,rh.rcv_organization_code
    ,rh.shipment_num
    ,rh.receipt_date
    ,rh.creation_date
    ,rh.process_flag process_flag_hd
    ,rl.process_flag process_flag_hl
    ,rl.line_id
    ,rl.shipment_line_num
    ,rl.item_number
    ,rl.quantity
    ,rl.container_number
    ,rl.subinv_code
    ,rl.shipment_num
    ,rl.shipment_line_num
    ,rl.rma_number
    ,(
      CASE 
        WHEN INSTR(lower(rh.error_message),'full received') >0 and rl.internal_req_number is null and rh.shipment_num is not null THEN 'ASN Fully Received'
        WHEN INSTR(lower(rh.error_message),'full received') >0 and rl.internal_req_number is not null THEN 'IRN Fully Received'
        WHEN INSTR(lower(rh.error_message),'incorrect data shipment_num is not provided') >0 THEN 'Blind Receipts'
        WHEN INSTR(lower(rh.error_message),'po release is closed') >0  THEN 'PO Release is CLOSED'
        WHEN INSTR(lower(rh.error_message),'closed') >0  THEN 'PO Closed'
        WHEN INSTR(lower(rh.error_message),'cancelled') >0 THEN 'PO Line Cancelled'   
        WHEN INSTR(lower(rh.error_message),'not in approved status') >0 THEN 'PO Not Approved'   
        WHEN INSTR(lower(rh.error_message),'invalid combination') >0 THEN 'PO Line Changed'  
        WHEN INSTR(lower(rh.error_message),'quantity provided') >0 THEN 'UOM or Quantity Tolarance Issue'  
        WHEN rh.transaction_type='RMA' AND rh.shipment_num is not null THEN 'RMA Closed/Line Split to multiple lines'
        WHEN rh.transaction_type='RMA' AND rh.shipment_num is null THEN 'RMA Order Missing/RMA Cancelled'           
        ELSE 'No Error MSG'
      END
    ) ERROR_GROUP
    ,rh.error_message
    ,rl.attribute13
    ,rl.attribute14
    ,rl.subinventory
    ,rl.po_num
    ,rl.rel_num
    ,rl.po_line_num
    ,rl.po_shipment_num
    ,rl.internal_req_number
    ,rl.int_req_line_num
    ,rl.source_reference_id
    ,rh.source_reference_id
    --,rsh.receipt_num
    --,rt.po_unit_price
    --,(rl.quantity*rt.po_unit_price) recipt_line_value
    --,ooh.order_number RMA
    --,rsh.shipment_header_id
    --,rsh.creation_date
    --,rsh.receipt_source_code
    --,rsh.shipment_num
    --,rsh.receipt_num
    --,rsl.*
    from
    APPS.GEPO_APL_RECEIPT_INBOUND_TBL rh
    ,APPS.GEPO_APL_RECEIPT_LINES_INB_TBL rl
    --,apps.rcv_shipment_headers rsh
    --,apps.rcv_shipment_lines rsl
    --,apps.oe_order_headers_all ooh
    where 1=1
    and rh.header_id=rl.header_id
    and rh.source_reference_id=rl.source_reference_id
    and rh.rcv_organization_code in ('JEF','RVR','NAP','WAL')
    --and rh.process_flag ='E'
    --and rh.transaction_type='RECEIPT'
    --and rh.creation_date BETWEEN TO_DATE('"""+fromDate+"""', 'DD-MON-YYYY hh24:mi:ss')
    --AND TO_DATE('"""+toDate+"""', 'DD-MON-YYYY hh24:mi:ss')
    and rh.creation_date >= TO_DATE('"""+fromDate+"""', 'DD-MON-YYYY hh24:mi:ss')
    -- and rl.item_number='225D8399P003'
    --and rh.shipment_num='226467'
    --and rl.shipment_line_num=11
    --and rsh.shipment_num(+)=rh.shipment_num
    --and rsl.line_num(+)=rl.shipment_line_num
    --and rsh.shipment_header_id=rsl.shipment_header_id(+)
    --and rsl.oe_order_line_id(+)=rl.shipment_line_num
    --and rh.process_flag!='N'
    order by rh.transaction_type
    ,rh.rcv_organization_code
    ,rh.shipment_num
    """)
    
    query_result2=("""
    SELECT
    concat(mp.organization_code,msi.segment1) key1,
    --concat(concat(mp.organization_code,msi.segment1),decode(mtt.transaction_type_name,'RMA Receipt',to_char(ooh.order_number),mmt.SHIPMENT_NUMBER)) KEY2,  - 29042024
    -- CASE WHEN mp.organization_code in ('NAP','WAL')  THEN concat(concat(mp.organization_code,msi.segment1),decode(mtt.transaction_type_name,'RMA Receipt',to_char(ooh.order_number),mmt.SHIPMENT_NUMBER))
    --      WHEN mp.organization_code in ('JEF', 'RVR') AND rt.attribute12 IS NOT NULL  THEN 
    --     ( SELECT Concat('UID.',trim(rl.container_number)) FROM APPS.GEPO_APL_RECEIPT_LINES_INB_TBL rl WHERE rl.line_id = rt.attribute12 )
    CASE WHEN mp.organization_code in ('NAP','WAL')  THEN concat(concat(mp.organization_code,msi.segment1),decode(mtt.transaction_type_name,'RMA Receipt',to_char(ooh.order_number),mmt.SHIPMENT_NUMBER))
         WHEN mp.organization_code in ('JEF', 'RVR') AND rt.attribute12 IS NOT NULL  THEN 
         ( SELECT concat('UID.',trim(rl.container_number)) FROM APPS.GEPO_APL_RECEIPT_LINES_INB_TBL rl WHERE rl.line_id = rt.attribute12 )
    END KEY2,
    concat(concat(concat(mp.organization_code,msi.segment1),decode(mtt.transaction_type_name,'RMA Receipt',to_char(ooh.order_number),mmt.SHIPMENT_NUMBER)),mmt.transaction_quantity) KEY3,
    mp.organization_code,
    msi.item_type item_type,
    msi.segment1 item_number,
    mmt.SHIPMENT_NUMBER,
    msi.inventory_item_status_code,
    mmt.source_line_id,
    mmt.transaction_id,
    mmt.transaction_type_id,
    mtt.transaction_type_name
    --,MTA.reference_account
    --,GCC.concatenated_segments
    ,mmt.transaction_date --TO_DATE('DD-MON-YYYY hh24:mi:ss')
    ,mmt.creation_date
    ,mmt.last_update_Date
    --,mmt.subinventory
    ,mmt.transaction_quantity
    --,mta.primary_quantity cnt
    --,mta.base_transaction_value
    ,mmt.last_updated_by
    ,mmt.created_by
    ,fu.user_name
    ,fu.description
    ,ooh.order_number RMA
    ,rt.subinventory
    ,rt.po_unit_price
    ,rt.quantity qty_received
    ,ooh1.order_number sales_order
    ,rt.attribute12   -- Unique column added to the extract to identity the record in EBS staging and Base table added on April 3rd 2024 by Krishna -- 515140482
    FROM
    apps.mtl_parameters mp,
    apps.mtl_system_items_b msi,
    apps.mtl_material_transactions mmt
    ,apps.rcv_transactions rt
    ,apps.oe_order_headers_all ooh
    ,apps.fnd_user fu
    ,apps.mtl_transaction_types mtt
    ,apps.wsh_delivery_details wdd
    ,apps.oe_order_lines_all ool
    ,apps.oe_order_headers_all ooh1
    --apps.GL_CODE_COMBINATIONS_KFV GCC
    --apps.MTL_TRANSACTION_ACCOUNTS MTA
    WHERE
    1=1
    AND mmt.transaction_type_id in (10,13,12,15,16,18,19,22,27,1005,99,59,76,69,120,61,40,42,4) --= -- --in (31, 41) --in (31, 41,44,35) -= 44 --= 33 --
    AND mp.organization_code in ('NAP','WAL','JEF', 'RVR') --IN --
    AND mmt.inventory_item_id=msi.inventory_item_id
    AND mmt.organization_id=msi.organization_id
    AND mp.organization_id =msi.organization_id
    AND mmt.rcv_transaction_id=rt.transaction_id(+)
    AND rt.oe_order_header_id=ooh.header_id(+)
    --and mmt.transaction_id = mta.transaction_id
    --and mta.reference_account = gcc.code_combination_id
    --AND mmt.organization_id=mta.organization_id
    AND mmt.TRANSACTION_ACTION_ID =mmt.TRANSACTION_ACTION_ID+0
    AND mmt.TRANSACTION_SOURCE_TYPE_ID=mmt.TRANSACTION_SOURCE_TYPE_ID+0
    --AND mmt.transaction_date BETWEEN TO_DATE('17-JAN-2023 00:00:00', 'DD-MON-YYYY hh24:mi:ss')
    --AND TO_DATE('17-JAN-2023 23:59:59','DD-MON-YYYY hh24:mi:ss')
    --AND mmt.transaction_date BETWEEN TO_DATE('10-JUN-2023 00:00:00', 'DD-MON-YYYY hh24:mi:ss')
    --AND TO_DATE('11-JUN-2023 23:59:59','DD-MON-YYYY hh24:mi:ss')
    AND mmt.last_update_date >= TO_DATE('"""+fromDate+"""', 'DD-MON-YYYY hh24:mi:ss')  and mmt.last_update_date <= TO_DATE('"""+toDate4+"""', 'DD-MON-YYYY hh24:mi:ss')
    AND mmt.created_by=fu.user_id
    AND wdd.delivery_Detail_id(+) = mmt.picking_line_id
    and ool.line_id(+) = wdd.source_line_id
    and ooh1.header_id(+)=ool.header_id
    --AND wdd.source_code = 'OE'
    --and wdd.oe_interfaced_flag = 'Y'
    --and wdd.released_status = 'Y'
    --and mmt.last_updated_by in (32198 7 153740
    --and mmt.created_by in (32198,15370)
    and mtt.transaction_type_id = mmt.transaction_type_id
    --AND MMT.SOURCE_LINE_ID IN (271099929,85119,85121,85108)
    --and mmt.created_by NOT in (32198,75078, 153740,73977)
    --and msi.segment1 ='234D1704P001'
    --and transaction_type_name='Sales order issue'
    --and mmt.SHIPMENT_NUMBER='98811647'
    order by mmt.transaction_id
    """)
    
    #query_result3=("""
    #SELECT   
    #     Concat(Concat(mp.organization_code,msi.segment1),rti.shipment_num) key2
    #    ,Concat(Concat(Concat(mp.organization_code,msi.segment1),rti.shipment_num),rti.quantity) key3
    #    ,Concat(mp.organization_code,msi.segment1) key1
    #    ,mp.organization_code  
    #    ,msi.segment1 item_number
    #    ,rti.transaction_type
    #    ,rti.transaction_date
    #    ,rti.processing_status_code
    #    ,rti.transaction_status_code
    #    ,rti.quantity
    #    ,rti.item_id
    #    ,rti.auto_transact_code
    #    ,rti.receipt_source_code
    #    ,rti.to_organization_id
    #    ,rti.source_document_code
    #    ,rti.po_unit_price
    #    ,rti.destination_type_code
    #    ,rti.subinventory
    #    ,rti.shipment_num
    #    ,rti.attribute12
    #    ,rti.attribute14
    #    ,rti.item_num
    #    ,rti.oe_order_num
    #    ,rti.group_id
    #  --  ,rti.*
    #    FROM       
    #    apps.RCV_TRANSACTIONS_INTERFACE rti,
    #    apps.mtl_parameters mp,
    #    apps.mtl_system_items_b msi
    #    WHERE      1=1
    #    --AND processing_status_code = 'PENDING'
    #    and rti.creation_date  >= TO_DATE('01-JAN-2023 00:00:00', 'DD-MON-YYYY hh24:mi:ss')
    #    AND mp.organization_code IN ( 'JEF', 'WAL', 'NAP', 'RVR')
    #    AND rti.item_id=msi.inventory_item_id
    #    AND rti.to_organization_id=msi.organization_id
    #    --and shipment_num='69895340'
    #    --and receipt_source_code='INTERNAL ORDER'
    #    AND mp.organization_id =rti.to_organization_id
    #""")
    
    query_result3=("""
        select
        Concat(rh1.rcv_organization_code,rl1.item_number) KEY1
        ,Concat(Concat(rh1.rcv_organization_code,rl1.item_number),decode(rh1.transaction_type,'RMA',rl1.attribute14,rh1.shipment_num)) KEY2
        ,Concat(Concat(Concat(rh1.rcv_organization_code,rl1.item_number),decode(rh1.transaction_type,'RMA',rl1.attribute14,rh1.shipment_num)),rl1.quantity) KEY3
        ,rh1.transaction_type
        ,rh1.rcv_organization_code WAREHOUSE
        ,rh1.shipment_num
        ,rh1.receipt_date
        ,rh1.creation_date
        ,rh1.last_update_date
        ,rl1.creation_date CREATION_DATE
        ,trunc(rl1.creation_date) STG_DATE
        ,rl1.last_update_date rl_last_update_date
        ,rl1.last_update_date-rl1.creation_date REPROCESS_TIME
        ,rh1.process_flag process_flag_hd
        ,rl1.process_flag PROCESS_FLAG_HL
        ,rl1.line_id
        ,rl1.shipment_line_num
        ,rl1.item_number
        ,rl1.quantity RECEIPT_QUANTITY
        ,rl1.container_number
        ,rl1.subinv_code
        ,rl1.shipment_num
        ,rl1.shipment_line_num
        ,rl1.rma_number
        ,(
          CASE 
            WHEN INSTR(lower(rh1.error_message),'full received') >0 and rl1.internal_req_number is null and rh1.shipment_num is not null THEN 'ASN Fully Received'
            WHEN INSTR(lower(rh1.error_message),'full received') >0 and rl1.internal_req_number is not null THEN 'IRN Fully Received'
            WHEN INSTR(lower(rh1.error_message),'incorrect data shipment_num is not provided') >0 THEN 'Blind Receipts'
            WHEN INSTR(lower(rh1.error_message),'po release is closed') >0  THEN 'PO Release is CLOSED'
            WHEN INSTR(lower(rh1.error_message),'closed') >0  THEN 'PO Closed'
            WHEN INSTR(lower(rh1.error_message),'cancelled') >0 THEN 'PO Line Cancelled'   
            WHEN INSTR(lower(rh1.error_message),'not in approved status') >0 THEN 'PO Not Approved'   
            WHEN INSTR(lower(rh1.error_message),'invalid combination') >0 THEN 'PO Line Changed'  
            WHEN INSTR(lower(rh1.error_message),'quantity provided') >0 THEN 'UOM or Quantity Tolarance Issue'  
            WHEN rh1.transaction_type='RMA' AND rh1.shipment_num is not null THEN 'RMA Closed/Line Split to multiple lines'
            WHEN rh1.transaction_type='RMA' AND rh1.shipment_num is null THEN 'RMA Order Missing/RMA Cancelled'           
            ELSE 'No Error MSG'
          END
        ) ERROR_GROUP
        ,rh1.error_message
        ,rl1.attribute13
        ,rl1.attribute14
        ,rl1.subinventory
        ,rl1.po_num
        ,rl1.rel_num
        ,rl1.po_line_num
        ,rl1.po_shipment_num
        ,rl1.internal_req_number
        ,rl1.int_req_line_num
        ,rl1.source_reference_id
        ,rh1.source_reference_id
        --,rsh1.receipt_num
        --,rt1.po_unit_price
        --,(rl1.quantity*rt.po_unit_price) recipt_line_value
        --,ooh1.order_number RMA
        --,rsh1.shipment_header_id
        --,rsh1.creation_date
        --,rsh1.receipt_source_code
        --,rsh1.shipment_num
        --,rsh1.receipt_num
        ,(
          CASE 
            WHEN rh1.transaction_type='RECEIPT' and rl1.internal_req_number is not null then 'IRN' 
            WHEN rh1.transaction_type='RECEIPT' and rl1.internal_req_number is null and rh1.shipment_num is not Null then 'ASN' 
            WHEN rh1.transaction_type='RECEIPT' and rl1.internal_req_number is null and rh1.shipment_num is Null and rl1.po_num is not null then 'PO' 
            WHEN rh1.transaction_type='RMA' THEN 'RMA'
            ELSE 'UNIDENTIFED'
          END
         ) RECEIPT_TYPE    
        ,null as MTL_CREATION_DATE  
        ,null as IN_EBS
        ,null as RUN_TIME
        ,null as NOT_IN_EBS
        ,null as EXCEEDS_1_HR
        ,1 as ERRORS_PENDING_TO_REPROCESS
        from
          APPS.GEPO_APL_RECEIPT_INBOUND_TBL rh1
          ,APPS.GEPO_APL_RECEIPT_LINES_INB_TBL rl1
          where 1=1
          and rh1.header_id=rl1.header_id
          and rh1.source_reference_id=rl1.source_reference_id
          and rh1.rcv_organization_code in ('JEF','RVR','NAP','WAL')
          --and rh1.process_flag !='P'
        and rh1.creation_date BETWEEN TO_DATE('"""+(fromDate3)+"""', 'DD-MON-YYYY hh24:mi:ss')
        AND TO_DATE('"""+(toDate3)+"""', 'DD-MON-YYYY hh24:mi:ss')
        ORDER BY STG_DATE,PROCESS_FLAG_HD
    """)



    query_result4=("""
        select 
        concat(TH.warehouse,rc.ITEM) KEY1,
        -- concat(TH.warehouse,rc.ITEM,(if(rh.receipt_id_type='RMA',left(rc.receipt_id,10),ifnull(rh.SOURCE_FAX_NUM,'')))) KEY2,
        -- Concat('UID.',trim(cast(RC.INTERNAL_REC_CONT_NUM AS CHAR))) KEY2,
        concat('UID.',trim(cast(RC.INTERNAL_REC_CONT_NUM AS CHAR))) KEY2,
        concat(TH.warehouse,rc.ITEM,(if(rh.receipt_id_type='RMA',left(rc.receipt_id,10),ifnull(rh.SOURCE_FAX_NUM,''))),round(th.quantity,0)) KEY3, 
        rc.receipt_id,
        TH.WAREHOUSE, 
        rh.SOURCE_FAX_NUM, 
        NULL ERP_ORDER_NUMBER,
        rc.INTERNAL_RECEIPT_NUM, 
        rc.ITEM, 
        rc.QUANTITY RECEIPT_QUANTITY, 
        th.ACTIVITY_DATE_TIME, 
        DATE_FORMAT(th.ACTIVITY_DATE_TIME, '%Y-%m-%d') WMS_ACTIVITY_DATE_TIME, 
        RC.RECEIPT_ID_TYPE RECEIPT_ID_TYPE,
        th.TRANSACTION_TYPE, 
        th.REFERENCE_TYPE, 
        th.REFERENCE_ID ,
        RC.INTERNAL_REC_CONT_NUM ,  -- Unique column added to the extract to identity the record in EBS staging and Base table added on April 3rd 2024 by Krishna -- 515140482
        NULL IN_STAGE,
        NULL IN_EBS,
        NULL NOT_IN_STAGE,
        NULL NOT_IN_EBS,
        NULL RUN_TIME,
        NULL EXCEEDS_1_HR,
        NULL STG_CREATION_DATE,
        NULL MTL_CREATION_DATE,
        1 TOTAL_COUNT,
        null REPROCESS_COUNT
        from receipt_container RC
        INNER JOIN TRANSACTION_HISTORY TH ON TH.CONTAINER_ID = RC.CONTAINER_ID
        INNER JOIN RECEIPT_HEADER RH  ON RC.INTERNAL_RECEIPT_NUM = RH.INTERNAL_RECEIPT_NUM
        where RC.UPLOAD_INTERFACE_BATCH is not null
        and STATUS >= 301
        and th.ACTIVITY_DATE_TIME between '"""+fromDate2+"""' and '"""+toDate2+"""'
        and RC.RECEIPT_ID_TYPE in ('RMA','ASN','REQ','PO')
        and TH.REFERENCE_TYPE in ('RMA','ASN','REQ','PO')
        and TH.DIRECTION ='FROM'
        and (TH.TRANSACTION_TYPE in (120, 130) OR (TH.TRANSACTION_TYPE = 80 ))  -- and TH.CONTAINER_ID like 'MT%')) -- Changed on 12th Oct 2023 by Mangesh to include Scrap RMA
        and TH.LOCATION in (select LOCATION from LOCATION l join GENERIC_CONFIG_DETAIL gcd on l.LOCATION_CLASS = gcd.IDENTIFIER where RECORD_TYPE='LOCCLASS' AND SYS1VALUE = 'N')
        order by th.ACTIVITY_DATE_TIME
    """)
    
    query_result5=("""
        SELECT 
        concat(rst.warehouse,rst.item) KEY1
        ,concat(concat(rst.warehouse,rst.item),SUBSTRING_INDEX(erp_order_number,'-',1)) KEY2
        ,concat(rst.warehouse,rst.item,SUBSTRING_INDEX(erp_order_number,'-',1),quantity) KEY3
        ,null receipt_id
        ,rst.warehouse WAREHOUSE
        ,NULL SOURCE_FAX_NUM
        ,left(erp_order_number,10) ERP_ORDER_NUMBER
        ,rst.internal_receipt_num INTERNAL_RECEIPT_NUM
        ,rst.item ITEM
        ,rst.quantity RECEIPT_QUANTITY
        ,rst.date_and_time ACTIVITY_DATE_TIME
        ,DATE_FORMAT(rst.date_and_time, '%Y-%m-%d') WMS_ACTIVITY_DATE_TIME 
        ,'REC' RECEIPT_ID_TYPE
        ,'RECEIPT' TRANSACTION_TYPE
        ,'REC' REFERENCE_TYPE 
        ,rst.transaction_id REFERENCE_ID
        ,rst.internal_receipt_cont_num INTERNAL_REC_CONT_NUM  -- Unique column added to the extract to identity the record in EBS staging and Base table added on April 3rd 2024 by Krishna -- 515140482
        ,NULL IN_STAGE
        ,NULL IN_EBS
        ,NULL NOT_IN_STAGE
        ,NULL NOT_IN_EBS
        ,NULL RUN_TIME
        ,NULL EXCEEDS_1_HR
        ,NULL STG_CREATION_DATE
        ,NULL MTL_CREATION_DATE
        ,1 TOTAL_COUNT
        ,null REPROCESS_COUNT
        FROM irms_data.receipts rst
        WHERE warehouse in ('NAP','WAL')
        and rst.date_and_time between '"""+fromDate2+"""' and '"""+toDate2+"""'
    """)

    df4=pd.read_sql_query(query_result4,conn)
    header=list(df4.columns)
    print(df4.columns)
    print(file7)
    df5=pd.read_sql_query(query_result5,conn)
    header=list(df5.columns)
    print(df5.columns)
    combined_df = pd.concat([df4, df5], ignore_index=True)
    writer = pd.ExcelWriter(file7, engine='openpyxl')
    combined_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    df=pd.read_sql_query(query_result3,connection)
    header=list(df.columns)
    print(file6)
    writer = pd.ExcelWriter(file6, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    df_pending = df
    
    df=pd.read_sql_query(query_result1,connection)
    header=list(df.columns)
    print(file1)
    writer = pd.ExcelWriter(file1, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    df11 = df[(df['PROCESS_FLAG_HD'] != 'P') & (df['PROCESS_FLAG_HL'] != 'P')]
    header=list(df11.columns)
    print(file11)
    writer = pd.ExcelWriter(file11, engine='openpyxl')
    df11.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    df=pd.read_sql_query(query_result2,connection)
    header=list(df.columns)
    print(file2)
    writer = pd.ExcelWriter(file2, engine='openpyxl')
    df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    #df=pd.read_sql_query(query_result3,connection)
    #header=list(df.columns)
    #df.to_csv('query_result3.csv',index=False,header=header)
    
    #plsql_df1=pd.read_csv('query_result1.csv')
    #plsql_df2=pd.read_csv('query_result2.csv')
    #key_column = 'KEY3'
    #with open('query_result1.csv','r') as first_file, open('query_result2.csv','r') as second_file:
    #    first_dict = {row[key_column]: row for row in csv.DictReader(first_file)}
    #    second_dict = {row[key_column]: row for row in csv.DictReader(second_file)}
    #    unmatched_keys = set(first_dict.keys()) - set(second_dict.keys())
    #    with open('Unmatched_records.csv','w', newline='') as outfile:
    #        writer = csv.DictWriter(outfile, fieldnames=first_dict[list(first_dict.keys())[0]].keys())
    #        writer.writeheader()
    #        for key in unmatched_keys:
    #            writer.writerow(first_dict[key])
    
    #file11='query_result1.csv'
    #key11= 'KEY2'
    
    #file12='query_result2.csv'
    #key12= 'KEY2'
    
    #data1 = {}
    #with open(file11,'r') as csvfile:
    #    reader = csv.DictReader(csvfile)
    #    for row in reader:
    #        key = row[key11]
    #        if key not in data1:
    #            data1[key] = []
    #        data1[key].append(row)
    
    #data2 = {}
    #with open(file12,'r') as csvfile:
    #    reader = csv.DictReader(csvfile)
    #    for row in reader:
    #        data2[row[key12]] = row
    
    #unmatched_records = []
    #for key in data1.keys():
    #    if key not in data2:
    #       unmatched_records.extend(data1[key])
    
    #output_file = 'unmatched.csv'
    #with open(output_file, 'w', newline='') as csvfile:
    #    if unmatched_records:
    #        fieldnames = unmatched_records[0].keys()
    #        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    #        writer.writeheader()
    #        writer.writerows(unmatched_records)
     
    #os.rename('query_result1.csv',file1)
    #os.rename('query_result2.csv',file2)
    #os.rename('Unmatched.csv',file3)
    #os.rename('query_result3.csv',file4)

    key1 = 'KEY2'
    key2 = 'KEY2'
    key3 = 'KEY2'
    
    # Use pandas to read the Excel file into a DataFrame
    df1 = pd.read_excel(file1, sheet_name='Sheet1')
    df2 = pd.read_excel(file2, sheet_name='Sheet1')
    df3 = pd.read_excel(file7, sheet_name='Sheet1')

    data1 = {}
    data2 = {}
    
    # Read query_result1.xlsx and store data in data1
    df1 = pd.read_excel(file1)
    for _, row in df1.iterrows():
        key = row[key1]
        if key not in data1:
            data1[key] = []
        data1[key].append(row.to_dict())
    
    # Read query_result2.xlsx and store data in data2
    df2 = pd.read_excel(file2)
    for _, row in df2.iterrows():
        data2[row[key2]] = row.to_dict()

    
    unmatched_records = []
    matched_records = []
    for key in data1.keys():
        if key not in data2:
            unmatched_records += data1[key]
        else:
            matched_records += data1[key]

    # Convert the lists of unmatched and matched records to DataFrames
    unmatched_records_df = pd.DataFrame(unmatched_records)
    matched_records_df = pd.DataFrame(matched_records)
   
    header=list(unmatched_records_df.columns)
    print(file3)
    writer = pd.ExcelWriter(file3, engine='openpyxl')
    unmatched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    header=list(matched_records_df.columns)
    print(file4)
    writer = pd.ExcelWriter(file4, engine='openpyxl')
    matched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
   

    # Find unmatched & matched records from scale to stage
    scale_unmatched_records = []
    scale_matched_records = []
    
    for index, row in df3.iterrows():
        if row[key3] not in df1[key1].values and row[key3] not in df2[key2].values:
            scale_unmatched_record = row.to_dict()
            scale_unmatched_record['RUN_TIME'] = pd.Timedelta(0)
            scale_unmatched_record['IN_STAGE'] = 'NOT_IN_STAGE'
            scale_unmatched_record['NOT_IN_STAGE'] = 1
            scale_unmatched_record['IN_EBS'] = 'NOT_IN_EBS'
            scale_unmatched_record['NOT_IN_EBS'] = 1
            scale_unmatched_record['EXCEEDS_1_HR'] = 0  
            scale_unmatched_records.append(scale_unmatched_record)
            
        elif row[key3] in df1[key1].values and row[key3] in df2[key2].values:
            scale_matching_row1 = df1[df1[key1] == row[key3]]
            scale_matching_row2 = df2[df2[key2] == row[key3]]
            scale_matched_record = row.to_dict()
            scale_matched_record['STG_CREATION_DATE'] = scale_matching_row1['CREATION_DATE'].values[0]
            scale_matched_record['MTL_CREATION_DATE'] = scale_matching_row2['CREATION_DATE'].values[0]
            #scale_matched_record['RUN_TIME'] = scale_matching_row2['CREATION_DATE'].values[0].astype('datetime64[ns]') - row['ACTIVITY_DATE_TIME']
            scale_matched_record['RUN_TIME'] = scale_matching_row2['CREATION_DATE'].values[0].astype('datetime64[ns]') - scale_matching_row1['CREATION_DATE'].values[0].astype('datetime64[ns]')
            scale_matched_record['IN_STAGE'] = 'IN_STAGE'
            scale_matched_record['NOT_IN_STAGE'] = 0
            scale_matched_record['IN_EBS'] = 'IN_EBS'
            scale_matched_record['NOT_IN_EBS'] = 0
            scale_matched_record['EXCEEDS_1_HR'] = int(scale_matched_record['RUN_TIME'] > pd.Timedelta(hours=2))
            scale_matched_records.append(scale_matched_record)
            
        elif row[key3] in df1[key1].values and row[key3] not in df2[key2].values:
            scale_matching_row1 = df1[df1[key1] == row[key3]]
            scale_unmatched_record = row.to_dict()
            scale_unmatched_record['STG_CREATION_DATE'] = scale_matching_row1['CREATION_DATE'].values[0]
            scale_unmatched_record['RUN_TIME'] = pd.Timedelta(0)
            scale_unmatched_record['IN_STAGE'] = 'IN_STAGE'
            scale_unmatched_record['NOT_IN_STAGE'] = 0
            scale_unmatched_record['IN_EBS'] = 'NOT_IN_EBS'
            scale_unmatched_record['NOT_IN_EBS'] = 1
            scale_unmatched_record['EXCEEDS_1_HR'] = 0
            scale_unmatched_records.append(scale_unmatched_record)
            
        elif row[key3] not in df1[key1].values and row[key3] in df2[key2].values:
            scale_matching_row2 = df2[df2[key2] == row[key3]]
            scale_unmatched_record = row.to_dict()
            scale_unmatched_record['MTL_CREATION_DATE'] = scale_matching_row2['CREATION_DATE'].values[0]
            scale_unmatched_record['RUN_TIME'] = scale_matching_row2['CREATION_DATE'].values[0].astype('datetime64[ns]') - row['ACTIVITY_DATE_TIME']
            scale_unmatched_record['IN_STAGE'] = 'NOT_IN_STAGE'
            scale_unmatched_record['NOT_IN_STAGE'] = 1
            scale_unmatched_record['IN_EBS'] = 'IN_EBS'
            scale_unmatched_record['NOT_IN_EBS'] = 0
            scale_unmatched_record['EXCEEDS_1_HR'] = int(scale_unmatched_record['RUN_TIME'] > pd.Timedelta(hours=2))
            scale_unmatched_records.append(scale_unmatched_record)

    # Convert the lists of unmatched and matched records to DataFrames
    scale_unmatched_records_df = pd.DataFrame(scale_unmatched_records)
    scale_matched_records_df = pd.DataFrame(scale_matched_records)
   
    header=list(scale_unmatched_records_df.columns)
    print(file8)
    writer = pd.ExcelWriter(file8, engine='openpyxl')
    scale_unmatched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    header=list(scale_matched_records_df.columns)
    print(file9)
    writer = pd.ExcelWriter(file9, engine='openpyxl')
    scale_matched_records_df.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()
    
    # Combine the scale_unmatched and matched records
    result_scale_df = pd.concat([scale_unmatched_records_df, scale_matched_records_df], axis=0, ignore_index=True)
    result_scale_df = result_scale_df.sort_values(by=['WMS_ACTIVITY_DATE_TIME'])

    # Create a new Excel file with the combined results
    with pd.ExcelWriter(file10, engine='xlsxwriter') as writer:
        # Write the result DataFrame to a new sheet
        result_scale_df.to_excel(writer, sheet_name='Sheet1', index=False)
    
        # Format the 'Unmatched' and 'Matched' columns
        worksheet = writer.sheets['Sheet1']
        scale_unmatched_cond_fmt = {
            'type': 'cell',
            'criteria': 'equal to',
            'value': '"Unmatched"',
            'format': writer.book.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
        }
        scale_matched_cond_fmt = {
            'type': 'cell',
            'criteria': 'equal to',
            'value': '"Matched"',
            'format': writer.book.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
        }
        worksheet.conditional_format('G2:G{}'.format(len(result_scale_df)+1), scale_unmatched_cond_fmt)
        worksheet.conditional_format('G2:G{}'.format(len(result_scale_df)+1), scale_matched_cond_fmt)

    # Reorder the columns in result_scale_df
    column_order = ['WMS_ACTIVITY_DATE_TIME', 'IN_STAGE', 'IN_EBS', 'REFERENCE_TYPE', 'WAREHOUSE', 'TOTAL_COUNT', 'NOT_IN_STAGE', 'NOT_IN_EBS', 'EXCEEDS_1_HR', 'RUN_TIME', 'RECEIPT_QUANTITY']
    result_scale_df_ordered = result_scale_df[column_order]
    
    # Create an OrderedDict with the desired column order
    aggfunc_dict = OrderedDict([
        ('TOTAL_COUNT', 'sum'),
        ('NOT_IN_STAGE', 'sum'),
        ('NOT_IN_EBS', 'sum'),
        ('EXCEEDS_1_HR', 'sum'),
        ('RUN_TIME', [np.max, np.mean]),
        ('RECEIPT_QUANTITY', 'sum')
    ])
    
    # Perform the pivot operation using the ordered columns
    pt1 = pd.pivot_table(result_scale_df_ordered.reset_index(),
                         index=['WMS_ACTIVITY_DATE_TIME', 'IN_STAGE', 'IN_EBS', 'REFERENCE_TYPE', 'WAREHOUSE'],
                         aggfunc=aggfunc_dict)
    # Reorder the columns based on the desired column order
    #pt1 = pt1.reindex(columns=list(aggfunc_dict.keys()))
    # Rename columns in MultiIndex
    pt1.columns = list(map('_'.join, pt1.columns.values))
    # Convert PivotTable to HTML table
    html_table4 = pt1.to_html()

 
    # Calculate the defect percentage for each date
    defect_percent = result_scale_df.groupby('WMS_ACTIVITY_DATE_TIME')[['NOT_IN_STAGE','NOT_IN_EBS', 'EXCEEDS_1_HR', 'TOTAL_COUNT']].sum()
 
    
    # Convert the 'NOT_IN_EBS' and 'RECEIPT_COUNT' columns to numeric data types
    defect_percent['NOT_IN_STAGE'] = pd.to_numeric(defect_percent['NOT_IN_STAGE'], errors='coerce')
    defect_percent['NOT_IN_EBS'] = pd.to_numeric(defect_percent['NOT_IN_EBS'], errors='coerce')
    defect_percent['EXCEEDS_1_HR'] = pd.to_numeric(defect_percent['EXCEEDS_1_HR'], errors='coerce')
    defect_percent['TOTAL_COUNT'] = pd.to_numeric(defect_percent['TOTAL_COUNT'], errors='coerce')
    
    # Drop any rows with missing or non-numeric values
    defect_percent.dropna(inplace=True)
    
    # Calculate the defect percentage for each date
    defect_percent['PERCENTAGE_NOT_IN_STAGE'] = defect_percent['NOT_IN_STAGE'] / defect_percent['TOTAL_COUNT'] * 100
    defect_percent['Errors < 1 %'] = defect_percent['NOT_IN_EBS'] / defect_percent['TOTAL_COUNT'] * 100
    defect_percent['PERCENTAGE_EXCEEDS_1_HR'] = defect_percent['EXCEEDS_1_HR'] / defect_percent['TOTAL_COUNT'] * 100
    
    result1 = (defect_percent['EXCEEDS_1_HR'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]
    result2 = (defect_percent['NOT_IN_EBS'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]
    result3 = (defect_percent['NOT_IN_STAGE'] / defect_percent['TOTAL_COUNT'] * 100).iloc[0]
    result1 = float(result1)
    result1 = round(result1,2)
    result2 = float(result2)
    result2 = round(result2,2)
    result3 = float(result3)
    result3 = round(result3,2)

    defect_percent = defect_percent.rename(columns={'EXCEEDS_1_HR': 'EXCEEDS_PROCESS_TIME', 'PERCENTAGE_EXCEEDS_1_HR': 'Exceeds Process Time'})
    defect_percent = defect_percent.reset_index().rename(columns={'WMS_ACTIVITY_DATE_TIME': 'WMS_ACTIVITY_DATE_TIME'})
    
    # Convert the defect percent data frame to an HTML table string
    html_table5 = defect_percent.to_html(index=True)
    
    #    defect_pending = df_pending.groupby('STG_DATE','WAREHOUSE').agg(ERRORS_PENDING_TO_REPROCESS=('ERRORS_PENDING_TO_REPROCESS', 'sum','REPROCESS_TIME', 'sum')).reset_index()
    defect_pending = df_pending.groupby(['STG_DATE', 'WAREHOUSE']).agg({
        'ERRORS_PENDING_TO_REPROCESS': 'sum',
        'REPROCESS_TIME': 'sum'
    }).reset_index().rename(columns={
        'ERRORS_PENDING_TO_REPROCESS': 'ERRORS_PENDING_TO_REPROCESS_SUM',
        'REPROCESS_TIME': 'REPROCESS_TIME_SUM'
    })
        
    defect_percent['Errors not resolved within 24 hours'] =  df_pending.groupby('PROCESS_FLAG_HD').agg(ERRORS_PENDING_TO_REPROCESS=('ERRORS_PENDING_TO_REPROCESS', 'sum')).reset_index().iloc[0]['ERRORS_PENDING_TO_REPROCESS']      
        
    #defect_pending['ERRORS_PENDING_TO_REPROCESS'] = pd.to_numeric(defect_pending['ERRORS_PENDING_TO_REPROCESS'], errors='coerce') 
    defect_pending.dropna(inplace=True)

    if df_pending.empty:
        result4 = 0
    else:
        result4 = df_pending.groupby('PROCESS_FLAG_HD').agg(ERRORS_PENDING_TO_REPROCESS=('ERRORS_PENDING_TO_REPROCESS', 'sum')).reset_index().iloc[0]['ERRORS_PENDING_TO_REPROCESS']
    
    # Convert the DataFrame to an HTML table
    #html_table3 = defect_pending.to_html(index=False)
    
    data3 = pd.DataFrame()  # Create an empty DataFrame data3

    df3 = pd.read_excel(file2)  # Read data from file2 into df3

    for _, row in df3.iterrows():
        if pd.isnull(row[key3]):
            data3 = pd.concat([data3, row.to_frame().T], ignore_index=True)



    header=list(data3.columns)
    print(file5)
    writer = pd.ExcelWriter(file5, engine='openpyxl')
    data3.to_excel(writer, index=False, header=header, sheet_name='Sheet1')
    writer.close()

    data3.rename(columns={"ITEM_NUMBER": "# OF RECEIPTS CREATED MANUALLY", "TRANSACTION_QUANTITY": "RECEIPT_QTY"}, inplace=True)

    # Create an OrderedDict with the desired column order
    aggfunc_dict3 = OrderedDict([('# OF RECEIPTS CREATED MANUALLY', 'count'),
        ('RECEIPT_QTY', 'sum')])

    pt3 = pd.pivot_table(data3.reset_index(),
                         index=['ORGANIZATION_CODE', 'TRANSACTION_TYPE_NAME', 'USER_NAME', 'DESCRIPTION'],
                         aggfunc=aggfunc_dict3)
    # Reorder the columns based on the desired column order
    #pt1 = pt1.reindex(columns=list(aggfunc_dict.keys()))
    # Rename columns in MultiIndex
    pt3.columns = list(map('_'.join, pt3.columns.values))


    # Convert PivotTable to HTML table
    html_table3 = pt3.to_html()    
    
    
    # Open the Excel file and create an ExcelWriter object
    #writer = pd.ExcelWriter(file5, engine='openpyxl')
    #writer.book = openpyxl.load_workbook(file5)
    
    # Write the pivot table to a new sheet in the existing workbook
    #pt.to_excel(writer, sheet_name='Pivot Table')
    
    # Save the changes and close the workbook
    #writer.close()
    #writer.close()
    
    # Load the existing Excel file into a pandas DataFrame
    df_KPI = pd.read_excel('Parts_KPIs.xlsx')
    
    # Add custom text column to the pivot table DataFrame
    text_value = 'Receipts'
    defect_percent.insert(0, 'Parts_KPI', text_value)
    
    # Convert pivot table rows into a DataFrame
    df_append = pd.DataFrame(defect_percent.to_records())

    # Set a breakpoint
    #import pdb
    #pdb.set_trace()
    
    # Iterate over rows in df_append
    matching = 0
    for index, row in df_KPI.iterrows():
        #fromDate4 = row['WMS_ACTIVITY_DATE_TIME']
        existing_record1 = (df_append['Parts_KPI'] == text_value) & (df_append['WMS_ACTIVITY_DATE_TIME'] == fromDate4)
        #existing_record2 = (row['Parts_KPI'] == text_value) & (datetime.strptime(row['WMS_ACTIVITY_DATE_TIME'], "%Y-%m-%d").strftime("%Y-%m-%d") == fromDate4)
        existing_record2 = (row['Parts_KPI'] == text_value) & (row['WMS_ACTIVITY_DATE_TIME'] == fromDate4)
        
       
        # Compare with index record in df_KPI
        if len(df_append[existing_record1]) > 0:
            df_KPI_record = df_KPI.loc[index]  # Get the record at the current index in df_KPI
            append_record = df_append.loc[existing_record1].iloc[0]  # Get the matching record in df_append
            
            # Compare the record values
            if df_KPI_record.equals(append_record):
                matching = 1
                # Perform any desired actions or logic for matching records
    
        # Compare with index record in df_KPI
        if existing_record2:
            #if row['Parts_KPI'] == text_value and row['WMS_ACTIVITY_DATE_TIME'].strftime("%Y-%m-%d").upper() == fromDate4:
            if row['Parts_KPI'] == text_value and row['WMS_ACTIVITY_DATE_TIME'] == fromDate4:
                # Replace existing record with df_append record
                df_KPI.loc[index] = df_append.loc[existing_record1].iloc[0]

                matching = 1
    
        # Perform other operations or comparisons as needed
    
    if matching < 1:
        # Append new record
        df_KPI = pd.concat([df_KPI, df_append], ignore_index=False)
            
    # Convert 'Exceeds Process Time' column to numeric
    #df_KPI['Exceeds Process Time'] = pd.to_numeric(df_KPI['Exceeds Process Time'], errors='raise')
    
    # Apply formatting to 'Exceeds Process Time' column
    #df_KPI['Exceeds Process Time'] = df_KPI['Exceeds Process Time'].apply(lambda x: "{:.1f}%".format(x) if pd.notnull(x) else "")    
 
    # Save the updated DataFrame to a new Excel file
    df_KPI.to_excel('Parts_KPIs.xlsx', index=False)


    file0=file3
    usecols=[0] 
    team = 'CI: Oracle EBS R12 Core - PO Direct'
    subject = "CI: Oracle EBS R12 Core - PO Direct - EBS Receipt Errors from EBS Stage Table_{}".format(unstruct_toDate.strftime("%Y-%m-%d-%H-%M-%S").upper())
    emailbodytext = 'Please find attached the Receipt Errors file from EBS Stage, We need your attention to reprocess and determine the action on which you are not able to reprocess the receipt in EBS, Request you to check and reprocess the error receipts by today EOD. Hopefully Autmated Incident ticket will get created as the CI added on the subject along with Ops DL.'
    #files = [file1,file2,file3,file4,file5,file6,file7,file8,file9,file10]
    files = [file8,file1,file11,file5]


    if not os.path.isfile(file3):
        print(f"Error: '{file3}' does not exist")
        summary = ""
        grand_total=""
    elif os.stat(file3).st_size == 0:
        print(f"Error: '{file3}' is empty")
        summary = ""
        grand_total=""
    else:
            
        # Proceed with reading the XLSX file
        df = pd.read_excel(file3)

        if not df.empty:
            # Perform summary
            summary = df.groupby('ERROR_GROUP').agg({'ERROR_GROUP': 'count', 'QUANTITY': 'sum'})    
            # Print the summary
            print(summary)
            
            # Calculate the grand total
            grand_total = pd.Series({'ERROR_GROUP': summary['ERROR_GROUP'].sum(), 'QUANTITY': summary['QUANTITY'].sum()}, name='Grand Total')
            
            # Print the grand total
            print(grand_total)
            
            # Write the summary and grand total to an XLSX file


            output_file = file3
            
            # Load existing Excel file
            existing_file = pd.ExcelFile(output_file)
            
            # Read the existing sheets into a dictionary of DataFrames
            sheets_dict = {}
            for sheet_name in existing_file.sheet_names:
                sheets_dict[sheet_name] = existing_file.parse(sheet_name)
            
            # Concatenate the summary and grand_total DataFrames
            sheets_dict['Summary'] = pd.concat([summary, grand_total], ignore_index=True)
            
            # Create a new Workbook object
            workbook = Workbook()# Delete the default "Sheet"
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)
            # Iterate over sheets and write each DataFrame
            for sheet_name, df_sheet in sheets_dict.items():
                if sheet_name in workbook.sheetnames:
                    # If sheet already exists, remove it from the workbook
                    workbook.remove(workbook[sheet_name])
            
                # Create a new sheet and write the DataFrame
                workbook.create_sheet(sheet_name)
                ws = workbook[sheet_name]
                for r in dataframe_to_rows(df_sheet, index=False, header=True):
                    ws.append(r)
            
            # Save the Workbook to the output file
            workbook.save(output_file)
            # Close the existing file
            existing_file.close()
            
            print(f'Summary appended to {output_file}')
        else:          
            summary = ""
            grand_total= ""
            print(f"The file {file3} is empty.")


    email_proc(sender_email, receiv_email, cc_email, passwd_email, unstruct_toDate, file0, usecols, team, attention, emailbodytext, files, subject,summary,grand_total,html_table3,html_table4,html_table5,result1, result2, result3, result4)
  
    # Iterate through files in the directory
    for filename in os.listdir(working_folder):
        if filename.startswith('EBS_Receipts') and filename.endswith('.xlsx'):
            file_path = os.path.join(working_folder, filename)
            os.remove(file_path)            
    # Iterate through files in the directory
    for filename in os.listdir(working_folder):
       if filename.startswith('WMS_Receipts') and filename.endswith('.xlsx'):
            file_path = os.path.join(working_folder, filename)
            os.remove(file_path)            

def email_proc(sender_email, receiv_email, cc_email, passwd_email, unstruct_toDate, file0, usecols, team, attention, emailbodytext, files, subject,summary,grand_total,html_table3,html_table4,html_table5,result1, result2, result3, result4):

    import pandas as pd
    from datetime import datetime, timedelta
    from email.message import EmailMessage
    import smtplib
    import os.path
    import ssl
    #from passwd import pswd
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from openpyxl import load_workbook
    
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiv_email
    msg['CC'] = cc_email
    msg['Subject'] = subject
    
    print(receiv_email)
    print(cc_email)
    
    #if not os.path.isfile(file0):
    #    print(f"Error: '{file0}' does not exist")
    #elif os.stat(file0).st_size == 0:
    #    print(f"Error: '{file0}' is empty")
    #else:
    #    # Read the CSV file into a pandas DataFrame
    #    df = pd.read_csv(file0)
        
    #    # Convert the pandas DataFrame to a CSV file
    #    df.to_csv(file0, index=False)
        
    #    # Read CSV file and extract desired rows and columns
    #    df = pd.read_csv(file0, usecols=usecols)
    #    nrows = df.shape[0]
        
    #    # Read CSV file again using nrows
    #    df = pd.read_csv(file0, usecols=usecols, nrows=nrows)
        
    #    # Convert DataFrame to HTML table
    #    html_table = df.to_html(index=False)
        
        # Convert DataFrame to styled HTML table
    #html_table = df.style \
    #    .set_properties(**{'border-collapse': 'collapse', 'border': '1px solid black'}) \
    #    .set_table_styles([
    #        {'selector': 'th', 'props': [('background-color', 'lightblue'), ('font-weight', 'bold'), ('border', '1px solid black')]},
    #        {'selector': 'tr:last-child', 'props': [('background-color', 'lightblue'), ('font-weight', 'bold'), ('border', '1px solid black')]}
    #    ]) \
    #    .set_table_attributes('border="1" class="dataframe table table-striped table-hover table-sm" style="text-align: center; font-family: Arial"') \
    #    .set_table_attributes([{'selector': 'thead tr', 'props': [('background-color', 'lightblue')]}]) \
    #    .set_table_attributes([{'selector': 'tbody tr:nth-child(1), tbody tr:nth-child(2)', 'props': [('background-color', 'lightblue')]}]) \
    #    .to_html()

    # Add HTML message to email body
    html_content = f"""
    <html>
        <body>
            <p> </p>
            <p>{team},</p>
            <p>Kind attention required from: OPS Team</p>
            <p> </p>
            <p>{emailbodytext}</p>
            <p> </p>
            <p>KPI For Receipts<p>
            <p>  1. Receipts should get processed in Oracel within 2 hours (All receipts, RMA, IRN, ASN, PO Receipts) (Scale to EBS) :-  <b>{result1} % Receipts are Processed after 2 hours</b> <p>
            <p>  2. Errors should be Less than 1% error                                      :-  <b>{result2} % Errors</b> <p>
            <p>  3. Errors should resolve within 24 hours                                   :-  <b>{result4} Record/s of Receipt is/are Pending to reprocess by Ops Team</b>  <p>
            <p>Parts Automation-Python</p>
        </body>
    </html>
    """

    if not os.path.isfile(file0):
        print(f"Error: '{file0}' does not exist")
        html_body = MIMEText(html_content , 'html')
    elif os.stat(file0).st_size == 0:
        print(f"Error: '{file0}' is empty")
        html_body = MIMEText(html_content , 'html')
    else:
        # Create an HTML table string from the summary DataFrame
        if len(summary) == 0:
            html_body = MIMEText(html_content + html_table4 + html_table3 + html_table5, 'html')
        else:    
            html_table = summary.to_html()
            grand_total = grand_total.to_frame()
            html_table1 = grand_total.to_html()
            #html_body = MIMEText(html_content + html_table + html_table1 + html_table4 + html_table5 + html_table3, 'html')
            html_body = MIMEText(html_content  + html_table4 + html_table5 + html_table3, 'html')
        #print(html_content)
        #print(html_table)
        
        # Add HTML table to email body
        #html_body = MIMEText(html_content + html_table + html_table1 + html_table4 + html_table3 + html_table5, 'html')

    msg.attach(html_body)
    
    # Attach multiple files to email
    files = files
    
    # List of paths to files
    for file in files:
        with open(file, 'rb') as f:
            file_content = f.read()
            # Create attachment and add to message
            attachment = MIMEApplication(file_content, Name=file)
            attachment['Content-Disposition'] = f'attachment; filename="{file}"'
            msg.attach(attachment)
    
    # Send email
    #with smtplib.SMTP('smtp.office365.com', 587) as smtp:
    #    context = ssl.create_default_context()
    #    smtp.ehlo()
    #    smtp.starttls(context=context)
    #    smtp.login(sender_email, passwd_email)
    #    smtp.sendmail(sender_email, receiv_email.split(','), msg.as_string())
    with smtplib.SMTP('smtp.office365.com', 587) as smtp:
        context = ssl.create_default_context()
        smtp.ehlo()
        smtp.starttls(context=context)
        smtp.login(sender_email, passwd_email)
        receiv_email_list = receiv_email.split(',')
        cc_email_list = cc_email.split(',')
        receiv_email_list.extend(cc_email_list)
        smtp.sendmail(sender_email, receiv_email_list, msg.as_string())        

        print("job has been succesfully executed ", datetime.now())
   
def main():
    import sys
    import datetime
    sub_proc1()
    
if __name__ == "__main__":
    main()